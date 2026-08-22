# -*- coding: utf-8 -*-
"""
Gera um dashboard HTML autocontido de CRÃ‰DITO DISPONÃVEL do BCMS
(UASG 160329 - OGU e 167329 - Fundo do ExÃ©rcito) e das OMDS da Ba Ap Log Ex
a partir do export do Tesouro Gerencial 'CRÃ‰DITO DISP.xlsx' publicado no Google Drive.

- Baixa o xlsx pÃºblico do Drive (ou usa --local para testar com arquivo local).
- LÃª a aba 'CRÃ‰DITO DISP' (cabeÃ§alho dinÃ¢mico), valida o layout por nome de coluna,
  filtra as OMDS e calcula KPIs, quebras por AÃ§Ã£o/ND e detalhe por NC.
- VALIDAÃ‡ÃƒO anti-falha: CrÃ©dito DisponÃ­vel tem de fechar com Recebido âˆ’ Concedido âˆ’ Empenhado.
- Acumula um snapshot diÃ¡rio em data/history.json (para o grÃ¡fico de tendÃªncia).
- Escreve site/index.html (self-contained: CSS Moderno + Google Fonts + SVG + Tabela + Excel) e site/data/history.json.

Design UI/UX & Neurodesign SÃªnior (ui-ux-cognitive-engineering.md):
- Design tokens semÃ¢nticos completos em HSL (Light & Dark mode com zero fadiga visual).
- Tipografia fluida (clamp), Google Fonts Inter + JetBrains Mono + Newsreader.
- EquaÃ§Ã£o matemÃ¡tica em chips visuais (Recebido âˆ’ Empenhado = DisponÃ­vel).
- PÃ³dio gamificado 3D (ðŸ¥‡, ðŸ¥ˆ, ðŸ¥‰) com benchmarking de execuÃ§Ã£o orÃ§amentÃ¡ria.
- Badges cromÃ¡ticos de idade de crÃ©dito em tela (â‰¤30d, 31â€“60d, >60d).
- ExportaÃ§Ã£o nativa para Excel (.xls formatado) com tipos de dados numÃ©ricos e toast feedback.
- Modal drill-down com fÃ­sica spring e backdrop blur.
- Conformidade estrita WCAG 2.2 AA/AAA, 60fps GPU e zero clichÃªs amadores.

Uso:
    python gerar_dashboard.py                 # baixa do Drive (FileId padrÃ£o / env DRIVE_FILE_ID)
    python gerar_dashboard.py --local X.xlsx  # usa arquivo local (teste)
    python gerar_dashboard.py --date 2026-07-14  # forÃ§a a data do snapshot (default: hoje)
"""
import os, sys, json, argparse, datetime, urllib.request, tempfile, html, math, re, shutil
import openpyxl

HDR_ROW, DATA_ROW = 8, 9
# Manifesto das OMDS da Ba Ap Log â€” cada OM = par OGU (16xxxx) + FEx (167xxx).
UNIDADES = [
    {"sigla": "BCMS",      "nome": "BatalhÃ£o Central de ManutenÃ§Ã£o e Suprimento", "ogu": "160329", "fex": "167329", "logo": "BCMS.png",    "accent": "#CE2B2B", "key": "BCMS"},
    {"sigla": "Ba Ap Log", "nome": "Base de Apoio LogÃ­stico do ExÃ©rcito",          "ogu": "160238", "fex": "167238", "logo": "BAAPLOG.png", "accent": "#D83030", "key": "BAAPLOG"},
    {"sigla": "D C Mun",   "nome": "DepÃ³sito Central de MuniÃ§Ã£o",                  "ogu": "160246", "fex": "167246", "logo": "DCMUN.png",   "accent": "#047CC0", "key": "DCMUN"},
    {"sigla": "BMSA",      "nome": "BMSA",                                         "ogu": "160304", "fex": "167304", "logo": "BMSA.png",    "accent": "#DB2819", "key": "BMSA"},
    {"sigla": "1Âº D Sup",  "nome": "1Âº DepÃ³sito de Suprimento",                    "ogu": "160307", "fex": "167307", "logo": "1DSUP.png",   "accent": "#DE2B30", "key": "DSUP1"},
    {"sigla": "ECT",       "nome": "ECT",                                          "ogu": "160321", "fex": "167321", "logo": "Ect.png",     "accent": "#B33338", "key": "ECT"},
]
def _par(u):  # par de UASGs (OGU, FEx) de uma unidade, no formato (cod, label)
    return [(u["ogu"], f'{u["sigla"]} Â· OGU'), (u["fex"], f'{u["sigla"]} Â· FEx')]

ALVOS = [p for u in UNIDADES for p in _par(u)]
FONTE_CURTA = {}
for _u in UNIDADES:
    FONTE_CURTA[_u["ogu"]] = "160"; FONTE_CURTA[_u["fex"]] = "167"
DEFAULT_FILE_ID = "1Jv546wpWQSFAlep3oLRAg29hVy86iJxJ"
HERE = os.path.dirname(os.path.abspath(__file__))
SITE = os.path.join(HERE, "site")
DATA = os.path.join(HERE, "data")
HISTFILE = os.path.join(DATA, "history.json")

# ---------------- leitura ----------------
def norm(s): return str(s).strip().upper() if s is not None else ""

def to_num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace("'", "")
    if s in ("", "-", "-9", "NAO SE APLICA", "NÃƒO SE APLICA"): return 0.0
    try: return float(s)
    except ValueError:
        try: return float(s.replace(".", "").replace(",", "."))
        except ValueError: return 0.0

def disp(v):
    s = "" if v is None else str(v).strip().replace("'", "")
    return "" if s in ("-9", "NAO SE APLICA", "NÃƒO SE APLICA") else s

def baixar(file_id):
    url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (dashboard-bcms)"})
    tmp = os.path.join(tempfile.gettempdir(), "credito_disp_download.xlsx")
    with urllib.request.urlopen(req, timeout=120) as r, open(tmp, "wb") as f:
        f.write(r.read())
    if os.path.getsize(tmp) < 1000:
        raise SystemExit("Download muito pequeno â€” verifique o compartilhamento pÃºblico do arquivo.")
    return tmp

# ---------------- ETL ----------------
def _dt_br(s):
    try:
        dd, mm, aa = str(s).split("/")
        return datetime.date(int(aa), int(mm), int(dd))
    except Exception:
        return None

def anotar_trocas_nd(res):
    """Identifica as TROCAS INTERNAS DE ND ("mudanÃ§a/alteraÃ§Ã£o de ND") e rastreia a NC de ORIGEM.

    Como a planilha nÃ£o traz o vÃ­nculo, ele Ã© deduzido da ESTRUTURA (validado em 100% dos
    261 casos de AGO/2026): uma troca de ND Ã© uma mesma NC que, dentro da MESMA AÃ§Ã£o+PI,
    lanÃ§a valores opostos que se anulam â€” sai da ND antiga (linha negativa) e entra na ND
    nova (linha positiva). A NC de origem Ã© entÃ£o a que trouxe o crÃ©dito para a cÃ©lula da
    ND antiga; se essa tambÃ©m for uma troca, sobe-se a cadeia atÃ© a DESCENTRALIZAÃ‡ÃƒO
    original (a que carrega o objeto real). Anota em cada linha positiva:
        L["nd_de"]     -> ND de onde o crÃ©dito veio
        L["nc_origem"] -> {nc, obj, dia, emit, op} da NC que originou o crÃ©dito
    """
    for cod, d in res.items():
        linhas = d["linhas"]
        if not linhas:
            continue
        por_nc = {}
        for L in linhas:
            if L["nc"]:
                por_nc.setdefault(L["nc"], []).append(L)
        # trocas[nc][(acao,pi)] = (nd_origem, nd_destino, valor)
        trocas = {}
        for nc, ls in por_nc.items():
            grupos = {}
            for L in ls:
                grupos.setdefault((L["acao"], L["pi"]), []).append(L)
            for k, gl in grupos.items():
                pos = [x for x in gl if x["cred"] > 0.005]
                neg = [x for x in gl if x["cred"] < -0.005]
                if pos and neg and abs(sum(x["cred"] for x in gl)) <= 0.05:
                    trocas.setdefault(nc, {})[k] = (neg[0]["nd"], pos[0]["nd"], pos[0]["cred"])
        if not trocas:
            continue
        # entradas de crÃ©dito por cÃ©lula (candidatas a origem)
        entradas = {}
        for L in linhas:
            if L["cred"] > 0.005 and L["nc"]:
                entradas.setdefault((L["acao"], L["pi"], L["nd"]), []).append(L)

        def rastrear(nc_alt, acao, pi, nd_org, valor, prof=0, visto=None):
            if visto is None:
                visto = set()
            if prof > 6 or nc_alt in visto:
                return None
            visto.add(nc_alt)
            dt_alt = None
            for L in por_nc.get(nc_alt, []):
                dt_alt = _dt_br(L["dia"])
                if dt_alt:
                    break
            cands = [L for L in entradas.get((acao, pi, nd_org), [])
                     if L["nc"] != nc_alt
                     and (_dt_br(L["dia"]) or datetime.date.min) <= (dt_alt or datetime.date.max)]
            if not cands:
                return None
            exatos = [c for c in cands if abs(c["cred"] - valor) < 0.05]
            cands = exatos or sorted(cands, key=lambda c: (_dt_br(c["dia"]) or datetime.date.min), reverse=True)
            orig = cands[0]
            t = trocas.get(orig["nc"], {}).get((acao, pi))
            if t:  # a origem tambÃ©m Ã© troca de ND -> sobe a cadeia atÃ© a descentralizaÃ§Ã£o real
                acima = rastrear(orig["nc"], acao, pi, t[0], t[2], prof + 1, visto)
                if acima:
                    return acima
            return orig

        for nc, mp in trocas.items():
            for (acao, pi), (nd_org, nd_dst, val) in mp.items():
                orig = rastrear(nc, acao, pi, nd_org, val)
                for L in por_nc.get(nc, []):
                    if L["acao"] == acao and L["pi"] == pi and L["cred"] > 0.005:
                        L["nd_de"] = nd_org
                        if orig:
                            L["nc_origem"] = {"nc": orig["nc"], "obj": orig["obj"], "dia": orig["dia"],
                                              "emit": orig["emit"], "op": orig["op"]}

def etl(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for nm in wb.sheetnames:
        if norm(nm).startswith("CREDITO DISP") or "CRÃ‰DITO DISP" in nm:
            ws = wb[nm]; break
    if ws is None:
        ws = wb.active
    hdr_row = None
    for r in range(1, 16):
        rowvals = {norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        if "CREDITO DISPONIVEL" in rowvals or "PROVISAO RECEBIDA" in rowvals:
            hdr_row = r; break
    if not hdr_row:
        raise SystemExit("NÃ£o encontrei o cabeÃ§alho (PROVISAO RECEBIDA / CREDITO DISPONIVEL) â€” a fonte no Drive pode ter mudado de formato.")
    hdr = {}
    for c in range(1, ws.max_column + 1):
        n = norm(ws.cell(hdr_row, c).value)
        if n and n not in hdr: hdr[n] = c
    def col(name, req=True):
        c = hdr.get(name)
        if not c and req:
            raise SystemExit(f"Coluna '{name}' nÃ£o encontrada (aba '{ws.title}') â€” a fonte mudou de layout.")
        return c
    C = dict(prov=col("PROVISAO RECEBIDA"), cred=col("CREDITO DISPONIVEL"),
             emp=col("DESPESAS EMPENHADAS"), liq=col("DESPESAS LIQUIDADAS"), pag=col("DESPESAS PAGAS"),
             conc=col("PROVISAO CONCEDIDA", req=False))
    # [FIX layout 2026-08] Os rÃ³tulos (AÃ§Ã£o/PI/ND/NC + descriÃ§Ãµes) ficam em linhas de
    # cabeÃ§alho ACIMA da linha de mÃ©tricas e a fonte JÃ MUDOU de posiÃ§Ãµes. Resolve por
    # NOME varrendo as linhas de cabeÃ§alho, com fallback p/ as posiÃ§Ãµes atuais conhecidas
    # (col 6/7/8/9/10/5/11/12/13/1). Antes liam posiÃ§Ãµes fixas ERRADAS (col 4/6/11/8/9/â€¦),
    # o que embaralhava os rÃ³tulos (AÃ§Ã£o lia o nome da UG, ND lia a descriÃ§Ã£o da NC) e
    # quebrava a soma por cÃ©lula â€” o empenho nÃ£o abatia o recebimento e o "em tela" inflava.
    hdrL = {}
    for rr in range(max(1, hdr_row - 4), hdr_row + 1):
        for c in range(1, ws.max_column + 1):
            nn = norm(ws.cell(rr, c).value)
            if nn and nn not in hdrL:
                hdrL[nn] = c
    def colL(name, fb):
        c = hdrL.get(name)
        return c if c else fb
    CA    = colL("ACAO GOVERNO", 6)           # cÃ³digo da AÃ§Ã£o de Governo
    CPI   = colL("PI", 7)                      # cÃ³digo do Plano Interno
    CPID  = CPI + 1                            # descriÃ§Ã£o do PI (coluna Ã  direita, sem cabeÃ§alho prÃ³prio)
    CND   = colL("NATUREZA DESPESA", 9)        # cÃ³digo da Natureza de Despesa
    CNDD  = CND + 1                            # descriÃ§Ã£o da ND
    CNC   = colL("NC", 5)                      # nÃºmero da Nota de CrÃ©dito
    COBJ  = colL("NC - DESCRICAO", 11)         # descriÃ§Ã£o/objeto da NC
    COP   = colL("NC - OPERACAO (TIPO)", 12)   # operaÃ§Ã£o (recebimento/detalhamento/anulaÃ§Ã£o)
    CDIA  = colL("NC - DIA EMISSAO", 13)       # data de emissÃ£o da NC
    CEMIT = colL("EMITENTE - UG", 1)           # UG emitente
    data_row = None
    for r in range(hdr_row + 1, min(hdr_row + 8, ws.max_row + 1)):
        if norm(ws.cell(r, 3).value).replace("'", "").isdigit():
            data_row = r; break
    if not data_row:
        data_row = hdr_row + 1
    periodo = None
    for r in range(max(1, hdr_row - 2), hdr_row + 3):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "").strip().upper()
            if re.match(r"^[A-Z]{3}/\d{4}$", v):
                periodo = v; break
        if periodo:
            break

    res = {c: {"prov": 0.0, "conc": 0.0, "cred": 0.0, "emp": 0.0, "liq": 0.0, "pag": 0.0,
               "nome": label, "linhas": [], "celulas": {}, "por_acao": {}, "por_nd": {}, "nd_nome": {}, "n": 0}
           for c, label in ALVOS}
    codigos_alvo = set(res.keys())
    ugs_presentes = set()

    for r in range(data_row, ws.max_row + 1):
        ug_raw = ws.cell(r, 3).value
        if ug_raw is None: continue
        ug = str(ug_raw).strip().replace("'", "")
        if not ug.isdigit(): continue
        ugs_presentes.add(ug)
        if ug not in codigos_alvo: continue

        prov = to_num(ws.cell(r, C["prov"]).value)
        conc = to_num(ws.cell(r, C["conc"]).value) if C["conc"] else 0.0
        cred = to_num(ws.cell(r, C["cred"]).value)
        emp  = to_num(ws.cell(r, C["emp"]).value)
        liq  = to_num(ws.cell(r, C["liq"]).value)
        pag  = to_num(ws.cell(r, C["pag"]).value)

        d = res[ug]
        d["prov"] += prov; d["conc"] += conc; d["cred"] += cred
        d["emp"]  += emp;  d["liq"]  += liq;  d["pag"]  += pag
        d["n"]    += 1

        acao     = disp(ws.cell(r, CA).value)
        pi       = disp(ws.cell(r, CPI).value)
        pi_nome  = disp(ws.cell(r, CPID).value)
        nd       = disp(ws.cell(r, CND).value)
        nd_nome  = disp(ws.cell(r, CNDD).value)
        nc       = disp(ws.cell(r, CNC).value)
        dia      = disp(ws.cell(r, CDIA).value)
        emit     = disp(ws.cell(r, CEMIT).value)
        op       = disp(ws.cell(r, COP).value)
        obj      = disp(ws.cell(r, COBJ).value)
        
        # fallback na busca de descriÃ§Ã£o caso a coluna 13 esteja vazia
        if not obj:
            for ci in (15, 16, 17, 18, 19, 20):
                if ci <= ws.max_column:
                    cand = disp(ws.cell(r, ci).value)
                    if len(cand) > 10 and not cand.replace(".", "").replace(",", "").replace("-", "").isdigit():
                        obj = cand; break

        d["linhas"].append(dict(
            acao=acao, pi=pi, pi_nome=pi_nome, nd=nd, nd_desc=nd_nome,
            nc=nc, dia=dia, emit=emit, op=op, obj=obj,
            prov=prov, conc=conc, cred=cred, emp=emp, liq=liq, pag=pag))

        if acao: d["por_acao"][acao] = d["por_acao"].get(acao, 0.0) + cred
        if nd:
            d["por_nd"][nd] = d["por_nd"].get(nd, 0.0) + cred
            if nd_nome and nd not in d["nd_nome"]: d["nd_nome"][nd] = nd_nome

        k_cel = (acao, pi, nd)
        if k_cel not in d["celulas"]:
            d["celulas"][k_cel] = {"acao": acao, "pi": pi, "pi_nome": pi_nome, "nd": nd, "nd_nome": nd_nome,
                                   "prov": 0.0, "conc": 0.0, "cred": 0.0, "emp": 0.0, "liq": 0.0, "pag": 0.0,
                                   "cpos": 0.0, "cneg": 0.0, "n_nc": 0, "ncs": []}
        cel = d["celulas"][k_cel]
        cel["prov"] += prov; cel["conc"] += conc; cel["cred"] += cred
        cel["emp"]  += emp;  cel["liq"]  += liq;  cel["pag"]  += pag
        # decompÃµe a col. CrÃ©dito DisponÃ­vel: movimentos + (recebido) e âˆ’ (consumido/empenho/anulaÃ§Ã£o)
        if cred >= 0: cel["cpos"] += cred
        else:         cel["cneg"] += -cred
        if nc:
            cel["n_nc"] += 1
            cel["ncs"].append((nc, dia, emit, op, cred, obj))

    total_linhas = sum(d["n"] for d in res.values())
    if total_linhas == 0:
        raise SystemExit(
            f"Nenhuma linha das UASGs da Ba Ap Log na aba '{ws.title}'. "
            f"UGs presentes: {', '.join(sorted(ugs_presentes)[:8]) or 'nenhuma'}. "
            "Verifique a planilha no Google Drive.")

    for d in res.values():
        for cel in d["celulas"].values():
            # Recebido (lÃ­q) âˆ’ Empenhado/Consumido = CrÃ©dito DisponÃ­vel (fecha por construÃ§Ã£o)
            cel["aloc"] = cel["cpos"]
            cel["emp"]  = cel["cneg"]

    anotar_trocas_nd(res)

    alertas = []
    for cod, d in res.items():
        if d["n"] == 0: continue
        saldo_calc = d["prov"] - d["conc"] - d["emp"]
        if abs(saldo_calc - d["cred"]) > 0.05:
            alertas.append(f"{d['nome']}: CrÃ©dito DisponÃ­vel ({d['cred']:.2f}) difere de Recebidoâˆ’Concedidoâˆ’Empenhado ({saldo_calc:.2f}).")
        if d["emp"] < -0.01:
            alertas.append(f"{d['nome']}: Empenhado negativo ({d['emp']:.2f}).")

    return res, periodo, alertas

# ---------------- histÃ³rico ----------------
def atualizar_historico(res, data_str):
    os.makedirs(DATA, exist_ok=True)
    hist = []
    if os.path.exists(HISTFILE):
        try:
            with open(HISTFILE, "r", encoding="utf-8") as f:
                hist = json.load(f)
        except Exception:
            hist = []
    hist = [h for h in hist if h.get("data") != data_str]
    snap = {"data": data_str}
    for cod, _ in ALVOS:
        snap[cod] = {k: round(res[cod][k], 2) for k in ("prov", "conc", "cred", "emp", "liq", "pag")}
    snap["total"] = {k: round(sum(res[c][k] for c, _ in ALVOS), 2) for k in ("prov", "conc", "cred", "emp", "liq", "pag")}
    hist.append(snap)
    hist.sort(key=lambda h: h.get("data", ""))
    with open(HISTFILE, "w", encoding="utf-8") as f:
        json.dump(hist, f, ensure_ascii=False, indent=1)
    return hist

# ---------------- formataÃ§Ã£o ----------------
def _fmt(v):
    s = f"{abs(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return s

def brl(v):
    return ("âˆ’R$Â " if v < 0 else "R$Â ") + _fmt(v)

def num(v):
    return ("âˆ’" if v < 0 else "") + _fmt(v)

def pct(a, b): return (100.0 * a / b) if b else 0.0

def abrev(v):
    a = abs(v); s = "âˆ’" if v < 0 else ""
    if a >= 1e6: return s + f"{a/1e6:.1f}".replace(".", ",") + " mi"
    if a >= 1e3: return s + f"{a/1e3:.0f} mil"
    return s + f"{a:.0f}"

def esc(s): return html.escape(str(s))

# ---------------- SVG ----------------
def _r(x, y, w, h, var, extra=""):
    return f'<rect x="{x:.1f}" y="{y:.1f}" width="{max(0,w):.1f}" height="{h:.1f}" style="fill:var(--{var})" {extra}/>'

def svg_util(recebido, empenhado, disponivel):
    """Barra empilhada de utilizaÃ§Ã£o: Recebido = Empenhado + DisponÃ­vel."""
    if recebido <= 0:
        return '<p class="vazio">Sem provisÃ£o recebida</p>'
    x0, x1, y, h, W, H = 4, 636, 40, 32, 640, 92
    plot = x1 - x0
    fe = max(0.0, min(1.0, empenhado / recebido))
    we = plot * fe
    wd = plot - we
    pe, pd = pct(empenhado, recebido), pct(disponivel, recebido)
    return f'''<svg viewBox="0 0 {W} {H}" class="svg" role="img" aria-label="UtilizaÃ§Ã£o do crÃ©dito recebido">
<title>ProvisÃ£o Recebida {brl(recebido)}: Empenhado {brl(empenhado)} ({pe:.1f}%) + CrÃ©dito DisponÃ­vel {brl(disponivel)} ({pd:.1f}%)</title>
<text x="{x0}" y="20" class="s-lbl">PROVISÃƒO RECEBIDA TOTAL Â· {esc(brl(recebido))}</text>
<line x1="{x0}" y1="26" x2="{x1}" y2="26" class="s-brk"/><line x1="{x0}" y1="26" x2="{x0}" y2="32" class="s-brk"/><line x1="{x1}" y1="26" x2="{x1}" y2="32" class="s-brk"/>
{_r(x0, y, we, h, "warning-main", 'rx="6"')}
{_r(x0+we, y, wd, h, "success-main", 'rx="6"')}
<line x1="{x0+we:.1f}" y1="{y}" x2="{x0+we:.1f}" y2="{y+h}" style="stroke:var(--bg-surface);stroke-width:2.5"/>
<text x="{x0+8}" y="{y+h+17}" class="s-seg">Empenhado {esc(brl(empenhado))} Â· {pe:.1f}%</text>
<text x="{x1-8}" y="{y+h+17}" text-anchor="end" class="s-seg s-seg-ok">CrÃ©dito DisponÃ­vel {esc(brl(disponivel))} Â· {pd:.1f}%</text>
</svg>'''

def svg_waterfall(recebido, empenhado, disponivel, mini=False):
    """Waterfall horizontal: Recebida â†’ (âˆ’)Empenhado â†’ (=)DisponÃ­vel."""
    if recebido <= 0:
        return '<p class="vazio">Sem dados</p>'
    if mini:
        W, H, xL, rh, gap, fs = 360, 118, 112, 26, 10, 10
        L1, L2, L3 = "Recebido", "(âˆ’) Empenhado", "(=) DisponÃ­vel"
    else:
        W, H, xL, rh, gap, fs = 720, 196, 190, 42, 16, 13
        L1, L2, L3 = "ProvisÃ£o Recebida", "(âˆ’) Empenhado", "(=) CrÃ©dito DisponÃ­vel"
    xR = W - 24
    plot = xR - xL
    R = recebido
    def X(v): return xL + (v / R) * plot
    y1 = 14; y2 = y1 + rh + gap; y3 = y2 + rh + gap
    xd = X(max(disponivel, 0))
    parts = [f'<svg viewBox="0 0 {W} {H}" class="svg" role="img" aria-label="ComposiÃ§Ã£o do crÃ©dito">',
             f'<title>ProvisÃ£o Recebida {brl(recebido)} menos Empenhado {brl(empenhado)} igual a CrÃ©dito DisponÃ­vel {brl(disponivel)}</title>',
             f'<desc>{esc(brl(recebido))} âˆ’ {esc(brl(empenhado))} = {esc(brl(disponivel))}</desc>']
    # linha 1 â€” Recebida
    parts.append(_r(xL, y1, plot, rh, "primary-600", 'rx="5"'))
    parts.append(f'<text x="{xL-10}" y="{y1+rh/2+4:.0f}" text-anchor="end" class="s-cat">{esc(L1)}</text>')
    parts.append(f'<text x="{xR-8}" y="{y1+rh/2+4:.0f}" text-anchor="end" class="s-val s-on">{esc(brl(recebido))}</text>')
    # linha 2 â€” Empenhado (flutuante)
    we = plot - (xd - xL)
    parts.append(_r(xd, y2, we, rh, "warning-main", 'rx="5"'))
    parts.append(f'<text x="{xL-10}" y="{y2+rh/2+4:.0f}" text-anchor="end" class="s-cat">{esc(L2)}</text>')
    parts.append(f'<text x="{xd+8:.1f}" y="{y2+rh/2+4:.0f}" class="s-val s-on">âˆ’{esc(brl(empenhado).replace("R$Â ","R$Â "))}</text>')
    # linha 3 â€” DisponÃ­vel
    parts.append(_r(xL, y3, xd - xL, rh, "success-main", 'rx="5"'))
    parts.append(f'<text x="{xL-10}" y="{y3+rh/2+4:.0f}" text-anchor="end" class="s-cat s-cat-ok">{esc(L3)}</text>')
    parts.append(f'<text x="{xd+8:.1f}" y="{y3+rh/2+4:.0f}" class="s-val s-ok">{esc(brl(disponivel))}</text>')
    # conectores tracejados
    parts.append(f'<line x1="{xR:.1f}" y1="{y1+rh}" x2="{xR:.1f}" y2="{y2}" class="s-conn"/>')
    parts.append(f'<line x1="{xd:.1f}" y1="{y2+rh}" x2="{xd:.1f}" y2="{y3}" class="s-conn"/>')
    parts.append('</svg>')
    return "".join(parts)

def svg_diverg(itens, titulo, max_itens=8):
    itens = [(k, v) for k, v in itens if round(v, 2) != 0]
    itens = sorted(itens, key=lambda x: abs(x[1]), reverse=True)
    resto = sum(v for _, v in itens[max_itens:])
    itens = itens[:max_itens]
    if round(resto, 2) != 0:
        itens.append(("Outras", resto))
    if not itens:
        return f'<div class="card chart"><div class="eyebrow">{esc(titulo)}</div><p class="vazio">Sem valores no perÃ­odo</p></div>'
    vmax = max(abs(v) for _, v in itens) or 1
    labW, rh, pad = 150, 32, 10
    zx = labW + 246
    half = 230
    W = zx + half + 70
    H = pad * 2 + rh * len(itens) + 16
    el = [f'<line x1="{zx}" y1="{pad}" x2="{zx}" y2="{pad+rh*len(itens):.0f}" class="s-zero"/>']
    for i, (k, v) in enumerate(itens):
        y = pad + i * rh
        w = abs(v) / vmax * half
        lbl = k if len(k) <= 24 else k[:23] + "â€¦"
        el.append(f'<text x="{labW-8}" y="{y+rh/2+4:.0f}" text-anchor="end" class="s-cat">{esc(lbl)}</text>')
        if v >= 0:
            el.append(_r(zx, y+5, w, rh-10, "success-main", f'rx="4"><title>{esc(k)}: {esc(brl(v))}</title></rect'.replace("/>", ">")))
            el.append(f'<text x="{zx+w+8:.1f}" y="{y+rh/2+4:.0f}" class="s-num s-ok">{esc(num(v))}</text>')
        else:
            el.append(_r(zx-w, y+5, w, rh-10, "danger-main", f'rx="4"><title>{esc(k)}: {esc(brl(v))}</title></rect'.replace("/>", ">")))
            el.append(f'<text x="{zx-w-8:.1f}" y="{y+rh/2+4:.0f}" text-anchor="end" class="s-num s-neg">{esc(num(v))}</text>')
    svg = f'<svg viewBox="0 0 {W} {H}" class="svg" role="img" aria-label="{esc(titulo)}">{"".join(el)}</svg>'
    return f'<div class="card chart"><div class="eyebrow">{esc(titulo)}</div>{svg}</div>'

def svg_funil(cod, d):
    emp, liq, pag = d["emp"], d["liq"], d["pag"]
    base = emp or 1
    W, rh, gap, xL = 340, 26, 12, 110
    plot = W - xL - 75
    H = 18 + 3 * (rh + gap)
    rows = [("Empenhado", emp, "stg1", ""), ("Liquidado", liq, "stg2", f"{pct(liq,emp):.0f}% do emp."),
            ("Pago", pag, "stg3", f"{pct(pag,liq):.0f}% do liq.")]
    el = []
    for i, (nome, val, cls, conv) in enumerate(rows):
        y = 12 + i * (rh + gap)
        w = max(4, abs(val) / base * plot)
        el.append(f'<text x="{xL-10}" y="{y+rh/2+4:.0f}" text-anchor="end" class="s-cat">{nome}</text>')
        el.append(f'<rect x="{xL}" y="{y}" width="{w:.1f}" height="{rh}" rx="4" style="fill:var(--{cls})"><title>{nome}: {esc(brl(val))}</title></rect>')
        el.append(f'<text x="{xL+w+8:.1f}" y="{y+rh/2+4:.0f}" class="s-num s-on2">{esc(num(val))}</text>')
        if conv:
            el.append(f'<text x="{W-4}" y="{y+rh/2+4:.0f}" text-anchor="end" class="s-conv">{conv}</text>')
    svg = f'<svg viewBox="0 0 {W} {H}" class="svg" role="img" aria-label="EstÃ¡gios {cod}">{"".join(el)}</svg>'
    return f'<div class="card chart"><div class="eyebrow">EstÃ¡gios da Despesa Â· {esc(cod)} {esc(FONTE_CURTA[cod])}</div>{svg}</div>'

def svg_tendencia(hist):
    wk, order = {}, []
    for h in hist:
        try:
            y, m, dd = (int(x) for x in h["data"].split("-"))
            key = datetime.date(y, m, dd).isocalendar()[:2]
        except Exception:
            key = h.get("data")
        if key not in wk:
            order.append(key)
        wk[key] = h
    semanas = [wk[k] for k in order]
    pts = [(h["data"], h["total"]["cred"]) for h in semanas]
    W, H, pl, pb, pt, pr = 720, 220, 70, 36, 18, 80
    pw, ph = W - pl - pr, H - pb - pt
    vals = [v for _, v in pts]
    vmin, vmax = min(vals + [0]), max(vals + [1])
    rng = (vmax - vmin) or 1
    n = len(pts)
    def X(i): return pl + (pw * (i / (n - 1)) if n > 1 else pw / 2)
    def Y(v): return pt + ph - ((v - vmin) / rng * ph)
    el = []
    for t in range(4):
        val = vmin + rng * t / 3; y = Y(val)
        el.append(f'<line x1="{pl}" y1="{y:.1f}" x2="{W-pr}" y2="{y:.1f}" class="s-grid"/>')
        el.append(f'<text x="{pl-10}" y="{y+4:.1f}" text-anchor="end" class="s-ax">{esc(abrev(val))}</text>')
    if n > 1:
        line = "M" + " L".join(f"{X(i):.1f},{Y(v):.1f}" for i, (_, v) in enumerate(pts))
        area = f"M{X(0):.1f},{Y(vmin):.1f} L" + " L".join(f"{X(i):.1f},{Y(v):.1f}" for i, (_, v) in enumerate(pts)) + f" L{X(n-1):.1f},{Y(vmin):.1f} Z"
        el.append(f'<path d="{area}" class="s-area"/>')
        el.append(f'<path d="{line}" class="s-line"/>')
    step = max(1, n // 6)
    for i, (dt, v) in enumerate(pts):
        show = (i % step == 0 or i == n - 1)
        el.append(f'<circle cx="{X(i):.1f}" cy="{Y(v):.1f}" r="4" class="s-dot"><title>{esc(dt)}: {esc(brl(v))}</title></circle>')
        if show:
            el.append(f'<text x="{X(i):.1f}" y="{H-pb+18}" text-anchor="middle" class="s-ax">{dt[8:10]}/{dt[5:7]}</text>')
    if n >= 1:
        dt, v = pts[-1]
        el.append(f'<text x="{X(n-1):.1f}" y="{Y(v)-12:.1f}" text-anchor="end" class="s-num s-ok">{esc(brl(v))}</text>')
    nota = "" if n > 1 else '<p class="vazio">A curva semanal se desenvolve a partir da 2Âª semana de histÃ³rico.</p>'
    svg = f'<svg viewBox="0 0 {W} {H}" class="svg" role="img" aria-label="TendÃªncia semanal do CrÃ©dito DisponÃ­vel">{"".join(el)}</svg>'
    return f'<div class="card chart wide"><div class="eyebrow">TendÃªncia HistÃ³rica Â· CrÃ©dito DisponÃ­vel Consolidado</div>{svg}{nota}</div>'

# ---------------- componentes HTML ----------------
def kpi_tile(label, valor, chip, cls):
    chip_html = f'<span class="chip">{esc(chip)}</span>' if chip else ""
    return (f'<div class="kpi kpi-{cls}"><div class="kpi-l">{esc(label)}</div>'
            f'<div class="kpi-v num">{esc(valor)}</div>{chip_html}</div>')

def uasg_card(cod, d):
    barp = pct(d["emp"], d["prov"])
    return (f'<div class="card uasg">'
            f'<div class="uasg-h"><span class="uasg-cod num">{esc(cod)}</span>'
            f'<span class="pill-fonte">{esc(FONTE_CURTA[cod])}</span>'
            f'<span class="uasg-nome">{esc(d["nome"].split("Â·")[1].strip())}</span></div>'
            f'<div class="uasg-disp"><span class="uasg-disp-l">CrÃ©dito DisponÃ­vel</span>'
            f'<span class="uasg-disp-v num">{esc(brl(d["cred"]))}</span></div>'
            f'<div class="uasg-eq num">{esc(brl(d["prov"]))} <i>âˆ’</i> {esc(brl(d["emp"]))}</div>'
            f'{svg_waterfall(d["prov"], d["emp"], d["cred"], mini=True)}'
            f'<div class="uasg-exec"><div class="exec-l"><span>Empenhado / Recebido</span><span class="num">{barp:.1f}%</span></div>'
            f'<div class="exec-track"><div class="exec-fill" style="width:{min(barp,100):.1f}%"></div></div></div>'
            f'</div>')

def tabela_html(tid, celulas, com_fonte, ativo):
    """RelaÃ§Ã£o de crÃ©dito EM TELA por cÃ©lula orÃ§amentÃ¡ria (saldo lÃ­quido positivo)."""
    cols = (["Fonte"] if com_fonte else []) + ["AÃ§Ã£o", "PI", "ND", "AplicaÃ§Ã£o", "Recebido (lÃ­q)", "Empenhado", "CrÃ©dito Disp."]
    ths = []
    for c in cols:
        numc = c in ("Recebido (lÃ­q)", "Empenhado", "CrÃ©dito Disp.")
        cls = ' class="num"' if numc else ''
        ths.append(f'<th{cls} tabindex="0" role="button" aria-sort="none" onclick="bcmsSort(this)" onkeydown="if(event.key==\'Enter\'||event.key==\' \'){{event.preventDefault();bcmsSort(this)}}">{esc(c)}<span class="sort"></span></th>')
    body = []
    tot = sum(c["cred"] for c in celulas)
    for c in celulas:
        fonte = f'<td><span class="pill-fonte">{esc(FONTE_CURTA.get(c.get("uasg",""),""))}</span></td>' if com_fonte else ''
        aplic = c.get("nd_nome") or c.get("pi_nome") or ""
        cid = esc(c.get("cid", ""))
        body.append(
            f'<tr class="cel-row" tabindex="0" role="button" data-cel="{cid}" title="Ver as notas de crÃ©dito desta cÃ©lula (descriÃ§Ã£o completa)" '
            f'onclick="bcmsCel(this)" onkeydown="if(event.key==\'Enter\'||event.key==\' \'){{event.preventDefault();bcmsCel(this)}}">'
            f'{fonte}<td>{esc(c["acao"])}</td><td class="mono2">{esc(c["pi"])}</td><td class="mono2">{esc(c["nd"])}</td>'
            f'<td class="obj" title="{esc(aplic)}">{esc(aplic[:60])}</td>'
            f'<td class="num" data-sort="{c["aloc"]:.2f}">{esc(brl(c["aloc"]))}</td>'
            f'<td class="num" data-sort="{c["emp"]:.2f}">{esc(brl(c["emp"]))}</td>'
            f'<td class="num anchor" data-sort="{c["cred"]:.2f}">{esc(brl(c["cred"]))}<i class="chev" aria-hidden="true">â€º</i></td></tr>')
    ncols = len(cols)
    tfoot = (f'<tfoot><tr><td colspan="{ncols-1}">TOTAL Â· {len(celulas)} cÃ©lula(s) com crÃ©dito em tela</td>'
             f'<td class="num anchor">{esc(brl(tot))}</td></tr></tfoot>')
    disp_style = "" if ativo else ' style="display:none"'
    return (f'<div class="tabpanel" id="{tid}" role="tabpanel"{disp_style}>'
            f'<div class="tbl-tools"><label class="visually-hidden" for="q-{tid}">Buscar</label>'
            f'<input type="search" id="q-{tid}" class="tbl-search" placeholder="Buscar por aÃ§Ã£o, PI, ND ou aplicaÃ§Ã£oâ€¦" oninput="bcmsSearch(this,\'{tid}\')">'
            f'<button type="button" class="btn-excel" onclick="bcmsExportTable(this,\'{tid}\',\'creditos_em_tela_{tid}\')" title="Baixar dados em planilha formatada para Excel"><span class="btn-excel-ic">ðŸ“Š</span> Exportar Excel</button>'
            f'<span class="tbl-count" id="cnt-{tid}" data-unit="cÃ©lula(s)" aria-live="polite">{len(celulas)} cÃ©lulas</span></div>'
            f'<div class="tbl-scroll"><table class="det"><thead><tr>{"".join(ths)}</tr></thead>'
            f'<tbody>{"".join(body)}</tbody>{tfoot}</table></div></div>')

# ---------------- pÃ¡gina da unidade ----------------
def conteudo_unidade(res, hist, data_str, periodo, u):
    ALVOS = _par(u); sfx = u["key"]
    tot = {k: sum(res[c][k] for c, _ in ALVOS) for k in ("prov", "conc", "cred", "emp", "liq", "pag", "n")}
    ger = datetime.datetime.now().strftime("%d/%m/%Y Ã s %H:%M")
    posicao = periodo if periodo else (data_str[8:10] + "/" + data_str[5:7] + "/" + data_str[0:4])

    delta_html = ""
    if len(hist) >= 2:
        dv = hist[-1]["total"]["cred"] - hist[-2]["total"]["cred"]
        if round(dv, 2) != 0:
            seta = "â–²" if dv > 0 else "â–¼"
            cls = "up" if dv > 0 else "down"
            delta_html = f'<div class="delta {cls}"><span>{seta}</span> {esc(num(dv))} <small>vs. dia anterior</small></div>'
        else:
            delta_html = '<div class="delta flat">Sem variaÃ§Ã£o vs. dia anterior</div>'
    else:
        delta_html = '<div class="delta flat">1Âº dia de histÃ³rico</div>'

    kpis = (kpi_tile("ProvisÃ£o Recebida", brl(tot["prov"]), "", "prov") +
            kpi_tile("Empenhado", brl(tot["emp"]), f'{pct(tot["emp"],tot["prov"]):.1f}% do recebido', "emp") +
            kpi_tile("Liquidado", brl(tot["liq"]), f'{pct(tot["liq"],tot["emp"]):.1f}% do empenhado', "liq") +
            kpi_tile("Pago", brl(tot["pag"]), f'{pct(tot["pag"],tot["liq"]):.1f}% do liquidado', "pag"))

    cards_uasg = "".join(uasg_card(cod, res[cod]) for cod, _ in ALVOS)

    acao, nd, nd_nome = {}, {}, {}
    for cod, _ in ALVOS:
        for k, v in res[cod]["por_acao"].items(): acao[k] = acao.get(k, 0) + v
        for k, v in res[cod]["por_nd"].items(): nd[k] = nd.get(k, 0) + v
        nd_nome.update(res[cod]["nd_nome"])
    nd_lbl = {(f'{k} {nd_nome.get(k,"")[:18]}').strip(): v for k, v in nd.items()}
    ch_acao = svg_diverg(acao.items(), "CrÃ©dito DisponÃ­vel por AÃ§Ã£o (R$)")
    ch_nd = svg_diverg(nd_lbl.items(), "CrÃ©dito DisponÃ­vel por Natureza de Despesa (R$)")
    funis = "".join(svg_funil(cod, res[cod]) for cod, _ in ALVOS)
    trend = svg_tendencia(hist)

    lin_idx = {}
    for cod, _ in ALVOS:
        for L in res[cod]["linhas"]:
            lin_idx.setdefault((cod, L["acao"], L["pi"], L["nd"]), []).append(L)
    celdata = {}
    cid_map = {}
    cid_seq = [0]
    def cid_for(uasg, c):
        key = (uasg, c["acao"], c["pi"], c["nd"])
        cid = cid_map.get(key)
        if cid is None:
            cid_seq[0] += 1
            cid = "%s_c%d" % (sfx, cid_seq[0])
            cid_map[key] = cid
            ncs = sorted(([L["nc"], L["op"], round(L["cred"], 2), L["obj"], L.get("emit", ""), L.get("dia", "")]
                          for L in lin_idx.get(key, [])), key=lambda x: -x[2])
            celdata[cid] = {"t": f'{c["acao"]} Â· PI {c["pi"]} Â· ND {c["nd"]}', "nome": c.get("nd_nome", ""),
                            "u": FONTE_CURTA.get(uasg, ""), "uasg": uasg,
                            "acao": c["acao"], "pi": c["pi"], "pinome": c.get("pi_nome", ""),
                            "nd": c["nd"], "ndnome": c.get("nd_nome", ""),
                            "r": round(c["aloc"], 2), "e": round(c["emp"], 2), "l": round(c.get("liq", 0.0), 2),
                            "p": round(c.get("pag", 0.0), 2), "d": round(c["cred"], 2), "ncs": ncs}
        c["cid"] = cid
        return cid
    def celulas_pos(cod):
        cl = [c for c in res[cod]["celulas"].values() if c["cred"] > 0.005]
        cl.sort(key=lambda x: x["cred"], reverse=True)
        for c in cl:
            cid_for(cod, c)
        return cl
    cons_cel = []
    for cod, _ in ALVOS:
        for c in celulas_pos(cod):
            cons_cel.append({**c, "uasg": cod, "cid": cid_for(cod, c)})
    cons_cel.sort(key=lambda x: x["cred"], reverse=True)
    ogu_c, fex_c = ALVOS[0][0], ALVOS[1][0]
    abas = (f'<button class="tab on" role="tab" aria-selected="true" tabindex="0" onclick="bcmsTab(this,\'tab-cons-{sfx}\')" onkeydown="bcmsTabKey(event,this)">Consolidado</button>'
            f'<button class="tab" role="tab" aria-selected="false" tabindex="-1" onclick="bcmsTab(this,\'tab-{ogu_c}\')" onkeydown="bcmsTabKey(event,this)">{ogu_c} Â· OGU</button>'
            f'<button class="tab" role="tab" aria-selected="false" tabindex="-1" onclick="bcmsTab(this,\'tab-{fex_c}\')" onkeydown="bcmsTabKey(event,this)">{fex_c} Â· FEx</button>')
    tabs = (tabela_html(f"tab-cons-{sfx}", cons_cel, True, True) +
            tabela_html(f"tab-{ogu_c}", celulas_pos(ogu_c), False, False) +
            tabela_html(f"tab-{fex_c}", celulas_pos(fex_c), False, False))

    movs = []
    for cod, _ in ALVOS:
        for L in res[cod]["linhas"]:
            if not L["nc"]:
                continue
            try:
                dd, mm, yy = L.get("dia", "").split("/")
                dt = datetime.date(int(yy), int(mm), int(dd))
            except Exception:
                dt = None
            movs.append({**L, "uasg": cod, "dt": dt})
    ncdata = {}
    for i, m in enumerate(movs, 1):
        nid = "%s_n%d" % (sfx, i)
        m["nid"] = nid
        ncdata[nid] = {"nc": m["nc"], "u": FONTE_CURTA.get(m["uasg"], ""), "acao": m["acao"],
                       "pi": m["pi"], "nd": m["nd"], "ndn": m.get("nd_desc", ""), "op": m["op"],
                       "dia": m.get("dia", ""), "val": round(m["cred"], 2), "obj": m["obj"]}
    datas = sorted({m["dt"] for m in movs if m["dt"]}, reverse=True)
    max_date = datas[0] if datas else None
    fmt_d = lambda d: d.strftime("%d/%m/%Y") if d else "â€”"
    daily = sorted([m for m in movs if m["dt"] == max_date] if max_date else [], key=lambda x: x["cred"], reverse=True)
    rec_d = sum(m["cred"] for m in daily if m["cred"] > 0)
    red_d = sum(m["cred"] for m in daily if m["cred"] < 0)

    def _th(h, numc, sortable=True):
        cls = ' class="num"' if numc else ''
        if not sortable:
            return f'<th{cls}>{esc(h)}</th>'
        return (f'<th{cls} tabindex="0" role="button" aria-sort="none" onclick="bcmsSort(this)" '
                f'onkeydown="if(event.key===\'Enter\'||event.key===\' \'){{event.preventDefault();bcmsSort(this)}}">{esc(h)}<span class="sort"></span></th>')

    def mov_row(m):
        neg = ' cell-neg' if m["cred"] < 0 else ''
        return (f'<tr class="cel-row" tabindex="0" role="button" data-nc="{esc(m["nid"])}" title="Detalhar a NC" '
                f'onclick="bcmsNC(this)" onkeydown="if(event.key===\'Enter\'||event.key===\' \'){{event.preventDefault();bcmsNC(this)}}">'
                f'<td class="mono2">{esc(m["nc"])}</td><td><span class="pill-fonte">{esc(FONTE_CURTA[m["uasg"]])}</span></td>'
                f'<td>{esc(m["acao"])}</td><td class="mono2">{esc(m["nd"])}</td>'
                f'<td class="obj" title="{esc(m["op"])}">{esc(m["op"][:28])}</td>'
                f'<td class="num anchor{neg}" data-sort="{m["cred"]:.2f}">{esc(brl(m["cred"]))}<i class="chev" aria-hidden="true">â€º</i></td></tr>')

    def mov_tabela(tid, lst):
        if not lst:
            return '<p class="vazio">Sem movimentaÃ§Ã£o de NC neste perÃ­odo.</p>'
        ths = _th("NC", False) + _th("Fonte", False) + _th("AÃ§Ã£o", False) + _th("ND", False) + _th("OperaÃ§Ã£o", False) + _th("Valor", True)
        body = "".join(mov_row(m) for m in lst)
        return (f'<div class="tbl-tools"><label class="visually-hidden" for="q-{tid}">Buscar</label>'
                f'<input type="search" id="q-{tid}" class="tbl-search" placeholder="Buscar por NC, aÃ§Ã£o, ND ou operaÃ§Ã£oâ€¦" oninput="bcmsSearch(this,\'{tid}\')">'
                f'<button type="button" class="btn-excel" onclick="bcmsExportTable(this,\'{tid}\',\'movimentacao_{tid}\')" title="Baixar movimentaÃ§Ã£o em Excel"><span class="btn-excel-ic">ðŸ“Š</span> Exportar Excel</button>'
                f'<span class="tbl-count" id="cnt-{tid}" data-unit="NC(s)" aria-live="polite">{len(lst)} NC(s)</span></div>'
                f'<div class="tbl-scroll" id="{tid}"><table class="det"><thead><tr>{ths}</tr></thead><tbody>{body}</tbody></table></div>')

    week_days = datas[:7]
    week_rows, tot_rec, tot_red, n_week = [], 0.0, 0.0, 0
    daydata = {}
    for idx, d in enumerate(week_days):
        dm = [m for m in movs if m["dt"] == d]
        rec = sum(m["cred"] for m in dm if m["cred"] > 0)
        red = sum(m["cred"] for m in dm if m["cred"] < 0)
        tot_rec += rec; tot_red += red; n_week += len(dm)
        dk = "%s_d%d" % (sfx, idx)
        daydata[dk] = {"d": fmt_d(d), "n": len(dm), "rec": round(rec, 2), "red": round(red, 2), "liq": round(rec + red, 2),
                       "ncs": sorted([[m["nc"], FONTE_CURTA[m["uasg"]], m["op"], round(m["cred"], 2), m["obj"]] for m in dm],
                                     key=lambda x: -x[3])}
        week_rows.append((d, len(dm), rec, red, rec + red, dk))

    def semana_tabela(rows):
        if not rows:
            return '<p class="vazio">Sem movimentaÃ§Ã£o na Ãºltima semana.</p>'
        heads = _th("Dia", False, False) + _th("NÂº NC", True, False) + _th("Recebido (+)", True, False) + _th("ReduÃ§Ãµes (âˆ’)", True, False) + _th("LÃ­quido", True, False)
        body = ""
        for d, n, rec, red, liq, dk in rows:
            body += (f'<tr class="cel-row" tabindex="0" role="button" data-day="{dk}" title="Ver as NCs deste dia" '
                     f'onclick="bcmsDay(this)" onkeydown="if(event.key===\'Enter\'||event.key===\' \'){{event.preventDefault();bcmsDay(this)}}">'
                     f'<td class="mono2">{fmt_d(d)}</td><td class="num">{n}</td>'
                     f'<td class="num col-pos">{esc(brl(rec))}</td><td class="num col-neg">{esc(brl(red))}</td>'
                     f'<td class="num anchor">{esc(brl(liq))}<i class="chev" aria-hidden="true">â€º</i></td></tr>')
        foot = (f'<tfoot><tr><td>TOTAL</td><td class="num">{n_week}</td>'
                f'<td class="num">{esc(brl(tot_rec))}</td><td class="num">{esc(brl(tot_red))}</td>'
                f'<td class="num anchor">{esc(brl(tot_rec + tot_red))}</td></tr></tfoot>')
        return f'<div class="tbl-scroll"><table class="det"><thead><tr>{heads}</tr></thead><tbody>{body}</tbody>{foot}</table></div>'

    asof = max_date or datetime.date.today()
    rec_por_cel = {}
    for cod, _ in ALVOS:
        for L in res[cod]["linhas"]:
            if L["cred"] > 0 and L.get("dia"):
                try:
                    dd, mm, yy = L["dia"].split("/"); dt = datetime.date(int(yy), int(mm), int(dd))
                except Exception:
                    dt = None
                if dt:
                    rec_por_cel.setdefault((cod, L["acao"], L["pi"], L["nd"]), []).append({**L, "dt": dt})

    emtela_ncs = []
    for cod, _ in ALVOS:
        for cl in celulas_pos(cod):
            recs = sorted(rec_por_cel.get((cod, cl["acao"], cl["pi"], cl["nd"]), []),
                          key=lambda L: L["dt"], reverse=True)
            restante = cl["cred"]
            for L in recs:
                if restante <= 0.005:
                    break
                val_nc = min(L["cred"], restante)
                restante -= val_nc
                dt = L["dt"]
                dias = (asof - dt).days if dt else None
                emtela_ncs.append({
                    "uasg": cod,
                    "fonte": FONTE_CURTA.get(cod, ""),
                    "nc": L.get("nc") or "(sem nÃºmero)",
                    "acao": cl["acao"],
                    "pi": cl["pi"],
                    "pi_nome": cl.get("pi_nome", ""),
                    "nd": cl["nd"],
                    "nd_nome": cl.get("nd_nome", ""),
                    "obj": L.get("obj", "") or "(sem descriÃ§Ã£o)",
                    "op": L.get("op", ""),
                    "emit": L.get("emit", ""),
                    "dia": L.get("dia", ""),
                    "dt": dt,
                    "dias": dias,
                    "cred": val_nc,
                    "cid": cl.get("cid", ""),
                    "nd_de": L.get("nd_de", ""),
                    "nc_origem": L.get("nc_origem"),
                })
            if restante > 0.005:
                emtela_ncs.append({
                    "uasg": cod,
                    "fonte": FONTE_CURTA.get(cod, ""),
                    "nc": "(Saldo em tela)",
                    "acao": cl["acao"],
                    "pi": cl["pi"],
                    "pi_nome": cl.get("pi_nome", ""),
                    "nd": cl["nd"],
                    "nd_nome": cl.get("nd_nome", ""),
                    "obj": f"Saldo remanescente em tela da cÃ©lula {cl['acao']} Â· PI {cl['pi']} Â· ND {cl['nd']}",
                    "op": "SALDO REMANESCENTE",
                    "emit": "",
                    "dia": "",
                    "dt": None,
                    "dias": None,
                    "cred": restante,
                    "cid": cl.get("cid", "")
                })

    emtela_ncs.sort(key=lambda x: (x["dias"] if x["dias"] is not None else -1, x["cred"]), reverse=True)
    # Dados por LINHA em tela (uma NC especÃ­fica), para o detalhamento nÃ£o misturar NCs
    # de objetos diferentes que apenas compartilham o mesmo PI.
    teladata = {}
    for _i, _c in enumerate(emtela_ncs, 1):
        _tid = "%s_t%d" % (sfx, _i)
        _c["tid"] = _tid
        _org = _c.get("nc_origem") or {}
        teladata[_tid] = {
            "nc": _c["nc"], "uasg": _c["uasg"], "u": _c.get("fonte", ""),
            "acao": _c["acao"], "pi": _c["pi"], "pinome": _c.get("pi_nome", ""),
            "nd": _c["nd"], "ndnome": _c.get("nd_nome", ""),
            "v": round(_c["cred"], 2), "dias": _c["dias"], "dia": _c.get("dia", ""),
            "obj": _c.get("obj", ""), "op": _c.get("op", ""), "emit": _c.get("emit", ""),
            "cid": _c.get("cid", ""), "nd_de": _c.get("nd_de", ""),
            "org": ({"nc": _org.get("nc", ""), "obj": _org.get("obj", ""), "dia": _org.get("dia", ""),
                     "emit": _org.get("emit", "")} if _org else None),
        }
    tot_emtela = sum(c["cred"] for c in emtela_ncs)
    idades = [c["dias"] for c in emtela_ncs if c["dias"] is not None]
    idade_media = round(sum(idades) / len(idades)) if idades else 0
    idade_max = max(idades) if idades else 0

    def et_row(c):
        aplic = c.get("nd_nome") or c.get("pi_nome") or ""
        acao_nd = f'{c["acao"]} Â· {c["nd"]}'
        dias = c["dias"]
        if dias is None:
            dcls, dtxt, dsort = "badge-age age-none", "â€”", -1
        else:
            if dias > 60:
                dcls = "badge-age age-red"
            elif dias > 30:
                dcls = "badge-age age-amber"
            else:
                dcls = "badge-age age-green"
            dtxt = f'{dias}d'
            dsort = dias
        refd = c["dt"].strftime("%d/%m/%y") if c["dt"] else "â€”"
        cid = esc(c.get("cid", ""))
        # Troca de ND: a descriÃ§Ã£o prÃ³pria ("MUDANÃ‡A DE ND", "DETALHAMENTOâ€¦") nÃ£o diz o objeto.
        # Mostra o objeto REAL da NC de origem, sinalizando a troca.
        org = c.get("nc_origem")
        if org and org.get("obj"):
            desc_completa = esc(org["obj"])
            selo = f'<span class="tag-nd" title="CrÃ©dito recebido por mudanÃ§a de ND (de {esc(c.get("nd_de",""))} para {esc(c["nd"])}) â€” objeto herdado da NC de origem {esc(org["nc"])}">â†ª ND {esc(c.get("nd_de",""))}</span> '
        else:
            desc_completa = esc(c.get("obj", ""))
            selo = ""
        desc_resumo = desc_completa[:118] + ("â€¦" if len(desc_completa) > 118 else "")
        desc_resumo = selo + desc_resumo
        # nÂº curto da NC (o completo fica no title e no modal): "160504â€¦2026NC401667" -> "NC 401667 Â· 160504"
        nc_full = str(c["nc"] or "")
        m_nc = re.search(r"NC(\d+)$", nc_full)
        if m_nc:
            nc_lbl = f'<span class="nc-num">NC {esc(m_nc.group(1))}</span> <span class="nc-ug">Â· {esc(nc_full[:6])}</span>'
        else:
            nc_lbl = f'<span class="nc-num">{esc(nc_full)}</span>'
        tid = esc(c.get("tid", ""))
        faixa = "none" if dias is None else ("r" if dias > 60 else ("a" if dias > 30 else "v"))
        return (f'<tr class="cel-row" tabindex="0" role="button" data-tela="{tid}" data-cel="{cid}" '
                f'data-fonte="{esc(c.get("fonte",""))}" data-acao="{esc(c["acao"])}" data-nd="{esc(c["nd"])}" '
                f'data-faixa="{faixa}" data-val="{c["cred"]:.2f}" title="Clique para ver o que estÃ¡ em tela desta NC" '
                f'onclick="bcmsTela(this)" onkeydown="if(event.key===\'Enter\'||event.key===\' \'){{event.preventDefault();bcmsTela(this)}}">'
                f'<td><span class="pill-fonte">{esc(c["fonte"])}</span></td>'
                f'<td class="mono2" title="{esc(nc_full)}">{nc_lbl}</td>'
                f'<td class="mono2">{esc(acao_nd)}</td>'
                f'<td class="obj" title="{desc_completa}" data-full-desc="{desc_completa}">{desc_resumo}</td>'
                f'<td class="mono2">{esc(refd)}</td>'
                f'<td class="num anchor" data-sort="{c["cred"]:.2f}">{esc(brl(c["cred"]))}</td>'
                f'<td class="num" data-sort="{dsort}"><span class="{dcls}">{dtxt}</span><i class="chev" aria-hidden="true">â€º</i></td></tr>')

    et_ths = (
        _th("Fonte", False) +
        _th("NC", False) +
        _th("AÃ§Ã£o Â· ND", False) +
        _th("DescriÃ§Ã£o do objeto da NC", False) +
        _th("Recebido em", False) +
        _th("CrÃ©dito em Tela", True) +
        _th("Idade", True)
    )

    # opÃ§Ãµes dos filtros â€” sÃ³ o que existe de fato nesta unidade
    _fontes = sorted({c.get("fonte", "") for c in emtela_ncs if c.get("fonte")})
    _acoes = sorted({c["acao"] for c in emtela_ncs if c["acao"]})
    _nds = sorted({c["nd"] for c in emtela_ncs if c["nd"]})
    _nd_lbl = {}
    for _c in emtela_ncs:
        if _c["nd"] and _c["nd"] not in _nd_lbl and _c.get("nd_nome"):
            _nd_lbl[_c["nd"]] = _c["nd_nome"]
    opt_fonte = '<option value="">Fonte: todas</option>' + "".join(
        f'<option value="{esc(_f)}">{esc(_f)}{" Â· OGU (exercÃ­cio corrente)" if _f == "160" else (" Â· FEx (exercÃ­cios anteriores)" if _f == "167" else "")}</option>' for _f in _fontes)
    opt_acao = '<option value="">AÃ§Ã£o: todas</option>' + "".join(
        f'<option value="{esc(_a)}">{esc(_a)}</option>' for _a in _acoes)
    opt_nd = '<option value="">ND: todas</option>' + "".join(
        f'<option value="{esc(_n)}">{esc(_n)}{esc(" â€” " + _nd_lbl[_n][:26]) if _nd_lbl.get(_n) else ""}</option>'
        for _n in _nds)
    _cb = "bcmsFiltra('%s')" % sfx
    _cbl = "bcmsLimpaFiltros('%s')" % sfx

    emtela_html = (
        '<section class="sec"><div class="eyebrow">CrÃ©ditos em tela â€” por Nota de CrÃ©dito (NC)</div>'
        '<div class="et-head">'
        f'<div class="et-kpi et-hero"><span>CrÃ©dito DisponÃ­vel em tela</span><b class="num">{esc(brl(tot_emtela))}</b></div>'
        f'<div class="et-kpi"><span>Notas de CrÃ©dito</span><b class="num">{len(emtela_ncs)}</b></div>'
        f'<div class="et-kpi"><span>Idade mÃ©dia</span><b class="num">{idade_media} dias</b></div>'
        f'<div class="et-kpi"><span>Mais antigo</span><b class="num">{idade_max} dias</b></div>'
        f'<div class="et-action"><button type="button" class="btn-excel btn-excel-lg" onclick="bcmsExportTable(this,\'tab-emtela-{sfx}\',\'creditos_em_tela_nc_{sfx}\')" title="Baixar relatÃ³rio detalhado de crÃ©ditos por NC em planilha Excel"><span class="btn-excel-ic">ðŸ“¥</span> Baixar RelatÃ³rio NC em Excel</button></div>'
        f'<div class="et-meta">PosiÃ§Ã£o {esc(posicao)}<br><span class="rh-delay">â± dados com ~24h de defasagem</span></div></div>'
        '<p class="sec-nota">RelaÃ§Ã£o dos <b>crÃ©ditos disponÃ­veis por Nota de CrÃ©dito (NC)</b> com descriÃ§Ã£o completa do objeto e <b>dias em tela</b> (desde o lanÃ§amento da NC). '
        '<b>Clique em uma linha</b> para abrir a ficha completa. Legenda de idade: <span class="badge-age age-green">â‰¤30d</span> recente Â· <span class="badge-age age-amber">31â€“60d</span> atenÃ§Ã£o Â· <span class="badge-age age-red">&gt;60d</span> crÃ­tico.</p>'
        f'<div class="tbl-tools"><label class="visually-hidden" for="q-tab-emtela-{sfx}">Buscar</label>'
        f'<input type="search" id="q-tab-emtela-{sfx}" class="tbl-search" placeholder="Buscar por NC, AÃ§Ã£o, ND ou palavras na descriÃ§Ã£o completaâ€¦" oninput="{_cb}">'
        f'<button type="button" class="btn-excel" onclick="bcmsExportTable(this,\'tab-emtela-{sfx}\',\'creditos_em_tela_nc_{sfx}\')" title="Exporta exatamente as linhas visÃ­veis, conforme os filtros aplicados"><span class="btn-excel-ic">ðŸ“Š</span> Exportar Excel</button>'
        f'<span class="tbl-count" id="cnt-tab-emtela-{sfx}" data-unit="NC(s) em tela" aria-live="polite">{len(emtela_ncs)} NC(s) em tela</span></div>'
        f'<div class="tbl-filtros" id="flt-{sfx}" role="group" aria-label="Filtros da lista de crÃ©ditos em tela">'
        '<span class="flt-lbl">Filtrar:</span>'
        f'<select class="flt" id="f-fonte-{sfx}" aria-label="Filtrar por fonte" onchange="{_cb}">{opt_fonte}</select>'
        f'<select class="flt" id="f-acao-{sfx}" aria-label="Filtrar por aÃ§Ã£o de governo" onchange="{_cb}">{opt_acao}</select>'
        f'<select class="flt" id="f-nd-{sfx}" aria-label="Filtrar por natureza de despesa" onchange="{_cb}">{opt_nd}</select>'
        f'<select class="flt" id="f-idade-{sfx}" aria-label="Filtrar por idade em tela" onchange="{_cb}">'
        '<option value="">Idade: todas</option><option value="v">â‰¤ 30 dias</option>'
        '<option value="a">31 a 60 dias</option><option value="r">&gt; 60 dias</option></select>'
        f'<button type="button" class="flt-limpa" onclick="{_cbl}" title="Limpar todos os filtros">âœ• Limpar</button>'
        f'<span class="flt-resumo" id="flt-res-{sfx}" aria-live="polite"></span></div>'
        f'<div class="tbl-scroll" id="tab-emtela-{sfx}"><table class="det det-compact"><thead><tr>{et_ths}</tr></thead>'
        f'<tbody>{"".join(et_row(c) for c in emtela_ncs)}</tbody>'
        f'<tfoot><tr><td colspan="5">TOTAL Â· {len(emtela_ncs)} Nota(s) de CrÃ©dito em tela</td><td class="num anchor">{esc(brl(tot_emtela))}</td><td>â€”</td></tr></tfoot></table></div></section>'
    )

    resumo_html = (
        emtela_html
        + f'<section class="sec"><div class="eyebrow">MovimentaÃ§Ã£o de NC â€” {fmt_d(max_date)} (dia anterior)</div>'
        f'<p class="sec-nota">Notas de crÃ©dito com lanÃ§amento em <b>{fmt_d(max_date)}</b> (Ãºltimo dia com movimento â€” dados com ~24h de defasagem): '
        f'<b>{len(daily)}</b> NC(s) Â· Recebido <b>{esc(brl(rec_d))}</b> Â· ReduÃ§Ãµes <b>{esc(brl(red_d))}</b> Â· LÃ­quido <b>{esc(brl(rec_d + red_d))}</b>. '
        'Clique em uma NC para detalhÃ¡-la.</p>'
        + mov_tabela(f"mov-dia-{sfx}", daily) + '</section>'
        '<section class="sec"><div class="eyebrow">Resumo semanal â€” Ãºltimos 7 dias com movimentaÃ§Ã£o</div>'
        '<p class="sec-nota">MovimentaÃ§Ã£o de NC por dia: recebimentos (+), reduÃ§Ãµes/anulaÃ§Ãµes (âˆ’) e lÃ­quido. <b>Clique em um dia</b> para ver as NCs daquele dia.</p>'
        + semana_tabela(week_rows) + '</section>'
    )

    hero_eq = (
        f'<div class="hero-eq-box"><span class="eq-tag">RECEBIDO</span><span class="eq-val num">{esc(brl(tot["prov"]))}</span></div>'
        f'<span class="hero-eq-sign">âˆ’</span>'
        f'<div class="hero-eq-box"><span class="eq-tag">EMPENHADO</span><span class="eq-val num eq-emp">{esc(brl(tot["emp"]))}</span></div>'
        f'<span class="hero-eq-sign">=</span>'
        f'<div class="hero-eq-box eq-highlight"><span class="eq-tag">DISPONÃVEL</span><span class="eq-val num eq-disp">{esc(brl(tot["cred"]))}</span></div>'
    )
    disp = "" if (u is UNIDADES[0]) else ' style="display:none"'
    frag = f"""<section class="unidade" data-key="{sfx}" data-sigla="{esc(u['sigla'])}"{disp}>
  <div class="toptabs" role="tablist" aria-label="VisÃµes do painel">
    <button class="toptab on" role="tab" aria-selected="true" onclick="bcmsView(this,'resumo')">ðŸ“‹ Resumo Executivo & CrÃ©ditos em Tela</button>
    <button class="toptab" role="tab" aria-selected="false" onclick="bcmsView(this,'completo')">ðŸ“Š Detalhamento Completo & GrÃ¡ficos</button>
  </div>
  <div class="view-resumo">
  {resumo_html}
  </div>
  <div class="view-completo" style="display:none">
  <section class="hero">
    <div class="hero-l">
      <div class="eyebrow">CrÃ©dito DisponÃ­vel Â· Consolidado {esc(u['sigla'])}</div>
      <div class="hero-num num">{esc(brl(tot["cred"]))}</div>
      <div class="hero-eq">{hero_eq}</div>
    </div>
    <div class="hero-r">
      {delta_html}
      {svg_util(tot["prov"], tot["emp"], tot["cred"])}
    </div>
  </section>

  <section class="sec">
    <div class="eyebrow">ComposiÃ§Ã£o Visual da Disponibilidade</div>
    <div class="card">{svg_waterfall(tot["prov"], tot["emp"], tot["cred"])}</div>
  </section>

  <section class="sec">
    <div class="eyebrow">Indicadores Globais de ExecuÃ§Ã£o</div>
    <div class="kpis">{kpis}</div>
  </section>

  <section class="sec">
    <div class="eyebrow">Desdobramento por Fonte de Recurso</div>
    <div class="grid2">{cards_uasg}</div>
  </section>

  <section class="sec">
    <div class="eyebrow">DistribuiÃ§Ã£o do CrÃ©dito DisponÃ­vel</div>
    <div class="grid2">{ch_acao}{ch_nd}</div>
  </section>

  <section class="sec">
    <div class="eyebrow">EstÃ¡gios da Despesa por Fonte</div>
    <div class="grid2">{funis}</div>
  </section>

  <section class="sec">{trend}</section>

  <section class="sec">
    <div class="eyebrow">CrÃ©dito DisponÃ­vel em tela â€” por cÃ©lula orÃ§amentÃ¡ria</div>
    <p class="sec-nota">Saldo <b>lÃ­quido</b> por cÃ©lula (AÃ§Ã£o Â· PI Â· ND): <b>Recebido (lÃ­q) âˆ’ Empenhado = CrÃ©dito DisponÃ­vel</b> em cada linha. "Recebido (lÃ­q)" jÃ¡ compensa alteraÃ§Ãµes de ND, detalhamentos e anulaÃ§Ãµes (a alteraÃ§Ã£o <b>nÃ£o Ã© somada</b> com a NC original). SÃ³ aparecem cÃ©lulas com saldo &gt; 0; a soma fecha com o total consolidado. <b>Clique em uma linha</b> para ver as notas de crÃ©dito da cÃ©lula com a descriÃ§Ã£o completa.</p>
    <div class="tabs" role="tablist" aria-label="CrÃ©dito em tela por UASG">{abas}</div>
    {tabs}
  </section>
  </div>
</section>"""
    return frag, celdata, ncdata, daydata, teladata

# ---------------- mÃ³dulo Ranking e Comparativo OMDS ----------------
def svg_comparativo_barras(u_stats, metric="cred", titulo="CrÃ©dito DisponÃ­vel por Unidade (R$)"):
    itens = [(u["sigla"], u[metric], u["accent"]) for u in u_stats if round(u[metric], 2) != 0]
    itens.sort(key=lambda x: x[1], reverse=True)
    if not itens:
        return f'<div class="card chart"><div class="eyebrow">{esc(titulo)}</div><p class="vazio">Sem valores no perÃ­odo</p></div>'
    vmax = max(x[1] for x in itens) or 1
    labW, rh, pad = 110, 36, 12
    zx = labW + 10
    plotW = 440
    W = zx + plotW + 140
    H = pad * 2 + rh * len(itens)
    el = []
    for i, (sigla, val, accent) in enumerate(itens):
        y = pad + i * rh
        w = max(4, (val / vmax) * plotW)
        el.append(f'<text x="{labW-8}" y="{y+rh/2+4:.0f}" text-anchor="end" class="s-cat" style="font-weight:700">{esc(sigla)}</text>')
        el.append(f'<rect x="{zx}" y="{y+4}" width="{w:.1f}" height="{rh-8}" rx="4" style="fill:{accent}"><title>{esc(sigla)}: {esc(brl(val))}</title></rect>')
        el.append(f'<text x="{zx+w+8:.1f}" y="{y+rh/2+4:.0f}" class="s-num s-on2" style="font-weight:700">{esc(brl(val))}</text>')
    svg = f'<svg viewBox="0 0 {W} {H}" class="svg" role="img" aria-label="{esc(titulo)}">{"".join(el)}</svg>'
    return f'<div class="card chart"><div class="eyebrow">{esc(titulo)}</div>{svg}</div>'

def svg_comparativo_exec(u_stats, media_cmd):
    itens = [(u["sigla"], u["exec_pct"], u["accent"], u["prov"], u["emp"]) for u in u_stats]
    itens.sort(key=lambda x: x[1], reverse=True)
    labW, rh, pad = 110, 36, 12
    zx = labW + 10
    plotW = 440
    W = zx + plotW + 90
    H = pad * 2 + rh * len(itens) + 24
    el = []
    x_media = zx + (min(100.0, media_cmd) / 100.0) * plotW
    el.append(f'<line x1="{x_media:.1f}" y1="{pad-4}" x2="{x_media:.1f}" y2="{H-pad-16}" stroke="var(--gold)" stroke-width="2" stroke-dasharray="4 3"/>')
    el.append(f'<text x="{x_media:.1f}" y="{H-pad-2}" text-anchor="middle" font-size="11" font-weight="700" fill="var(--gold)">MÃ©dia do Comando: {media_cmd:.1f}%</text>')
    for i, (sigla, pct_v, accent, prov, emp) in enumerate(itens):
        y = pad + i * rh
        w = max(4, (min(100.0, pct_v) / 100.0) * plotW)
        el.append(f'<text x="{labW-8}" y="{y+rh/2+4:.0f}" text-anchor="end" class="s-cat" style="font-weight:700">{esc(sigla)}</text>')
        el.append(f'<rect x="{zx}" y="{y+4}" width="{plotW}" height="{rh-8}" rx="4" fill="var(--track)"/>')
        el.append(f'<rect x="{zx}" y="{y+4}" width="{w:.1f}" height="{rh-8}" rx="4" style="fill:{accent}"><title>{esc(sigla)}: {pct_v:.1f}% empenhado ({esc(brl(emp))} de {esc(brl(prov))})</title></rect>')
        el.append(f'<text x="{zx+w+8:.1f}" y="{y+rh/2+4:.0f}" class="s-num" font-weight="700">{pct_v:.1f}%</text>')
    svg = f'<svg viewBox="0 0 {W} {H}" class="svg" role="img" aria-label="Taxa de ExecuÃ§Ã£o OrÃ§amentÃ¡ria por OMDS">{"".join(el)}</svg>'
    return f'<div class="card chart"><div class="eyebrow">Taxa de ExecuÃ§Ã£o OrÃ§amentÃ¡ria (% Empenhado / Recebido)</div>{svg}</div>'

def secao_comparativo_omds(res, hist, data_str, periodo):
    u_stats = []
    for u in UNIDADES:
        alvos_u = _par(u)
        tot_prov = sum(res[c]["prov"] for c, _ in alvos_u)
        tot_conc = sum(res[c]["conc"] for c, _ in alvos_u)
        tot_emp = sum(res[c]["emp"] for c, _ in alvos_u)
        tot_liq = sum(res[c]["liq"] for c, _ in alvos_u)
        tot_pag = sum(res[c]["pag"] for c, _ in alvos_u)
        tot_cred = sum(res[c]["cred"] for c, _ in alvos_u)
        n_cel = sum(len([c for c in res[cod]["celulas"].values() if c["cred"] > 0.005]) for cod, _ in alvos_u)
        exec_pct = pct(tot_emp, tot_prov)
        liq_pct = pct(tot_liq, tot_emp)
        pag_pct = pct(tot_pag, tot_liq)
        u_stats.append({
            "key": u["key"], "sigla": u["sigla"], "nome": u["nome"], "logo": u["logo"],
            "accent": u["accent"], "ogu": u["ogu"], "fex": u["fex"],
            "prov": tot_prov, "conc": tot_conc, "emp": tot_emp, "liq": tot_liq,
            "pag": tot_pag, "cred": tot_cred, "exec_pct": exec_pct,
            "liq_pct": liq_pct, "pag_pct": pag_pct, "n_cel": n_cel
        })
    
    cmd_prov = sum(x["prov"] for x in u_stats)
    cmd_emp = sum(x["emp"] for x in u_stats)
    cmd_liq = sum(x["liq"] for x in u_stats)
    cmd_pag = sum(x["pag"] for x in u_stats)
    cmd_cred = sum(x["cred"] for x in u_stats)
    cmd_n_cel = sum(x["n_cel"] for x in u_stats)
    cmd_exec_pct = pct(cmd_emp, cmd_prov)
    cmd_liq_pct = pct(cmd_liq, cmd_emp)
    cmd_pag_pct = pct(cmd_pag, cmd_liq)
    
    rank_exec = sorted(u_stats, key=lambda x: x["exec_pct"], reverse=True)
    
    podio_order = []
    if len(rank_exec) >= 2:
        podio_order.append((rank_exec[1], 2, "ðŸ¥ˆ 2Âº Lugar", "silver"))
    if len(rank_exec) >= 1:
        podio_order.append((rank_exec[0], 1, "ðŸ¥‡ 1Âº Lugar", "gold"))
    if len(rank_exec) >= 3:
        podio_order.append((rank_exec[2], 3, "ðŸ¥‰ 3Âº Lugar", "bronze"))
    
    podio_cards = []
    for u, pos, badge, cls in podio_order:
        podio_cards.append(
            f'<div class="podium-step podium-{cls}" onclick="trocaOMDSPorKey(\'{u["key"]}\')" title="Clique para abrir o painel detalhado de {esc(u["sigla"])}">'
            f'<div class="podium-badge">{badge}</div>'
            f'<div class="podium-avatar-wrap"><img src="assets/logos/{u["logo"]}" alt="{esc(u["sigla"])}" class="podium-logo" onerror="this.style.display=\'none\'"></div>'
            f'<div class="podium-sigla">{esc(u["sigla"])}</div>'
            f'<div class="podium-nome">{esc(u["nome"])}</div>'
            f'<div class="podium-stat-pill"><span class="stat-l">ExecuÃ§Ã£o</span><b class="stat-v num">{u["exec_pct"]:.1f}%</b></div>'
            f'<div class="podium-substat">DisponÃ­vel: <span class="num">{esc(brl(u["cred"]))}</span></div>'
            f'<button type="button" class="podium-btn" onclick="event.stopPropagation();trocaOMDSPorKey(\'{u["key"]}\')">Acessar Unidade â€º</button>'
            f'</div>'
        )
    
    kpis_cmd = (
        kpi_tile("ProvisÃ£o Recebida (Comando)", brl(cmd_prov), "6 OMDS", "prov") +
        kpi_tile("Empenhado (Comando)", brl(cmd_emp), f"{cmd_exec_pct:.1f}% de execuÃ§Ã£o", "emp") +
        kpi_tile("Liquidado (Comando)", brl(cmd_liq), f"{cmd_liq_pct:.1f}% do empenhado", "liq") +
        kpi_tile("CrÃ©dito DisponÃ­vel", brl(cmd_cred), f"{cmd_n_cel} cÃ©lulas ativas", "pag")
    )
    
    ch_cred = svg_comparativo_barras(u_stats, "cred", "CrÃ©dito DisponÃ­vel por Unidade (R$)")
    ch_exec = svg_comparativo_exec(u_stats, cmd_exec_pct)
    
    def _th_r(h, numc):
        cls = ' class="num"' if numc else ''
        return (f'<th{cls} tabindex="0" role="button" aria-sort="none" onclick="bcmsSort(this)" '
                f'onkeydown="if(event.key===\'Enter\'||event.key===\' \'){{event.preventDefault();bcmsSort(this)}}">{esc(h)}<span class="sort"></span></th>')

    ths = (
        _th_r("Pos.", False) +
        _th_r("OrganizaÃ§Ã£o Militar (OMDS)", False) +
        _th_r("UASGs", False) +
        _th_r("ProvisÃ£o Recebida", True) +
        _th_r("Empenhado", True) +
        _th_r("% ExecuÃ§Ã£o", True) +
        _th_r("CrÃ©dito DisponÃ­vel", True) +
        _th_r("Liquidado", True) +
        _th_r("% LiquidaÃ§Ã£o", True) +
        _th_r("Pago", True) +
        _th_r("CÃ©lulas", True) +
        _th_r("AÃ§Ã£o", False)
    )
    
    body_rows = []
    for pos, u in enumerate(rank_exec, 1):
        medalha = "ðŸ¥‡ 1Âº" if pos == 1 else ("ðŸ¥ˆ 2Âº" if pos == 2 else ("ðŸ¥‰ 3Âº" if pos == 3 else f"{pos}Âº"))
        bar_w = min(100.0, u["exec_pct"])
        body_rows.append(
            f'<tr class="cel-row" tabindex="0" role="button" onclick="trocaOMDSPorKey(\'{u["key"]}\')" '
            f'title="Clique para ir ao painel do {esc(u["sigla"])}" onkeydown="if(event.key===\'Enter\'||event.key===\' \'){{event.preventDefault();trocaOMDSPorKey(\'{u["key"]}\')}}">'
            f'<td class="mono2" style="font-weight:700">{medalha}</td>'
            f'<td><div class="tbl-om-cell"><img src="assets/logos/{u["logo"]}" alt="" class="tbl-om-logo" onerror="this.style.display=\'none\'"><b>{esc(u["sigla"])}</b> <span class="tbl-om-sub">{esc(u["nome"])}</span></div></td>'
            f'<td class="mono2">{esc(u["ogu"])} / {esc(u["fex"])}</td>'
            f'<td class="num" data-sort="{u["prov"]:.2f}">{esc(brl(u["prov"]))}</td>'
            f'<td class="num" data-sort="{u["emp"]:.2f}">{esc(brl(u["emp"]))}</td>'
            f'<td class="num" data-sort="{u["exec_pct"]:.2f}"><div class="tbl-pct-cell"><span style="font-weight:700">{u["exec_pct"]:.1f}%</span><div class="mini-track"><div class="mini-fill" style="width:{bar_w:.1f}%;background:{u["accent"]}"></div></div></div></td>'
            f'<td class="num anchor col-pos" data-sort="{u["cred"]:.2f}">{esc(brl(u["cred"]))}</td>'
            f'<td class="num" data-sort="{u["liq"]:.2f}">{esc(brl(u["liq"]))}</td>'
            f'<td class="num" data-sort="{u["liq_pct"]:.2f}">{u["liq_pct"]:.1f}%</td>'
            f'<td class="num" data-sort="{u["pag"]:.2f}">{esc(brl(u["pag"]))}</td>'
            f'<td class="num" data-sort="{u["n_cel"]}">{u["n_cel"]}</td>'
            f'<td><button type="button" class="tbl-action-btn" onclick="event.stopPropagation();trocaOMDSPorKey(\'{u["key"]}\')">Abrir â€º</button></td>'
            f'</tr>'
        )
    
    tfoot_tbl = (
        f'<tfoot><tr>'
        f'<td colspan="3"><b>TOTAL CONSOLIDADO DO COMANDO (6 OMDS)</b></td>'
        f'<td class="num"><b>{esc(brl(cmd_prov))}</b></td>'
        f'<td class="num"><b>{esc(brl(cmd_emp))}</b></td>'
        f'<td class="num"><b>{cmd_exec_pct:.1f}%</b></td>'
        f'<td class="num anchor col-pos"><b>{esc(brl(cmd_cred))}</b></td>'
        f'<td class="num"><b>{esc(brl(cmd_liq))}</b></td>'
        f'<td class="num"><b>{cmd_liq_pct:.1f}%</b></td>'
        f'<td class="num"><b>{esc(brl(cmd_pag))}</b></td>'
        f'<td class="num"><b>{cmd_n_cel}</b></td>'
        f'<td>â€”</td>'
        f'</tr></tfoot>'
    )
    
    tabela_ranking_html = (
        f'<div class="tbl-tools">'
        f'<label class="visually-hidden" for="q-tab-ranking-det">Buscar</label>'
        f'<input type="search" id="q-tab-ranking-det" class="tbl-search" placeholder="Buscar no comparativo por OMDS, UASG..." oninput="bcmsSearch(this,\'tab-ranking-det\')">'
        f'<button type="button" class="btn-excel btn-excel-lg" onclick="bcmsExportTable(this,\'tab-ranking-det\',\'ranking_comparativo_omds\')" title="Baixar comparativo completo das OMDS em planilha formatada para Excel"><span class="btn-excel-ic">ðŸ“Š</span> Exportar Planilha Excel</button>'
        f'<span class="tbl-count" id="cnt-tab-ranking-det" data-unit="unidades" aria-live="polite">6 unidades</span>'
        f'</div>'
        f'<div class="tbl-scroll" id="tab-ranking-det"><table class="det"><thead><tr>{ths}</tr></thead><tbody>{"".join(body_rows)}</tbody>{tfoot_tbl}</table></div>'
    )
    
    hero_eq_cmd = (
        f'<div class="hero-eq-box"><span class="eq-tag">PROVISÃƒO TOTAL</span><span class="eq-val num">{esc(brl(cmd_prov))}</span></div>'
        f'<span class="hero-eq-sign">âˆ’</span>'
        f'<div class="hero-eq-box"><span class="eq-tag">EMPENHADO TOTAL</span><span class="eq-val num eq-emp">{esc(brl(cmd_emp))}</span></div>'
        f'<span class="hero-eq-sign">=</span>'
        f'<div class="hero-eq-box eq-highlight"><span class="eq-tag">DISPONÃVEL COMANDO</span><span class="eq-val num eq-disp">{esc(brl(cmd_cred))}</span></div>'
    )
    
    frag = f"""<section class="unidade unidade-ranking" data-key="RANKING" data-sigla="Comando" style="display:none">
  <div class="ranking-header-card">
    <div class="rh-tag">ðŸ† BENCHMARKING ORÃ‡AMENTÃRIO & FINANCEIRO</div>
    <h2 class="rh-title">Ranking & Comparativo Consolidado das OMDS</h2>
    <p class="rh-desc">VisÃ£o executiva integrada das 6 OrganizaÃ§Ãµes Militares Diretamente Subordinadas da Base de Apoio LogÃ­stico do ExÃ©rcito. Acompanhe os indicadores de desempenho, taxa de execuÃ§Ã£o orÃ§amentÃ¡ria (% Empenhado) e crÃ©ditos em tela.</p>
  </div>

  <section class="hero hero-cmd">
    <div class="hero-l">
      <div class="eyebrow">CrÃ©dito DisponÃ­vel Â· Consolidado do Comando (6 OMDS)</div>
      <div class="hero-num num">{esc(brl(cmd_cred))}</div>
      <div class="hero-eq">{hero_eq_cmd}</div>
    </div>
    <div class="hero-r">
      <div class="delta flat">Consolidado das 12 UASGs (OGU + FEx)</div>
      {svg_util(cmd_prov, cmd_emp, cmd_cred)}
    </div>
  </section>

  <section class="sec">
    <div class="eyebrow">Indicadores Globais do Comando</div>
    <div class="kpis">{kpis_cmd}</div>
  </section>

  <section class="sec">
    <div class="eyebrow">ðŸ† PÃ³dio de EficiÃªncia OrÃ§amentÃ¡ria (% Empenhado / Recebido)</div>
    <p class="sec-nota">Destaque para as organizaÃ§Ãµes com maior percentual de execuÃ§Ã£o das dotaÃ§Ãµes orÃ§amentÃ¡rias recebidas no exercÃ­cio. Clique em uma unidade para detalhar.</p>
    <div class="podium-wrap">{"".join(podio_cards)}</div>
  </section>

  <section class="sec">
    <div class="eyebrow">Comparativos Visuais entre as Unidades</div>
    <div class="grid2">{ch_exec}{ch_cred}</div>
  </section>

  <section class="sec">
    <div class="eyebrow">Tabela Comparativa e Benchmarking Completo</div>
    <p class="sec-nota">RelaÃ§Ã£o completa de todas as OMDS com dados consolidados de OGU e FEx. Ordene por qualquer coluna ou clique no botÃ£o para exportar para Excel (.xlsx).</p>
    {tabela_ranking_html}
  </section>
</section>"""
    return frag

# ---------------- shell da pÃ¡gina (multi-OMDS) ----------------
def montar_pagina(res, hist, data_str, periodo=None, alertas=None):
    frags, CEL, NCD, DAY, TELA = [], {}, {}, {}, {}
    for u in UNIDADES:
        hist_u = [{"data": h.get("data"),
                   "total": {"cred": round(h.get(u["ogu"], {}).get("cred", 0.0)
                                           + h.get(u["fex"], {}).get("cred", 0.0), 2)}}
                  for h in hist]
        frag, cel, ncd, day, tela = conteudo_unidade(res, hist_u, data_str, periodo, u)
        frags.append(frag); CEL.update(cel); NCD.update(ncd); DAY.update(day); TELA.update(tela)
    
    ranking_frag = secao_comparativo_omds(res, hist, data_str, periodo)
    frags.append(ranking_frag)

    banner = ""
    if alertas:
        itens = "".join(f"<li>{esc(a)}</li>" for a in alertas)
        banner = f'<div class="banner" role="alert"><b>âš  VerificaÃ§Ã£o de ConsistÃªncia:</b><ul>{itens}</ul></div>'
    omds = "".join(
        f'<button class="omds{" on" if i == 0 else ""}" data-key="{u["key"]}" aria-current="{"true" if i == 0 else "false"}" '
        f'title="{esc(u["nome"])}" onclick="trocaOMDS(this)">'
        f'<img src="assets/logos/{u["logo"]}" alt="" loading="lazy" onerror="this.style.display=\'none\'">'
        f'<span>{esc(u["sigla"])}</span></button>'
        for i, u in enumerate(UNIDADES))
    omds += (
        '<button class="omds omds-rank" data-key="RANKING" aria-current="false" '
        'title="Ranking e Comparativo entre todas as Unidades da Base de Apoio LogÃ­stico" onclick="trocaOMDS(this)">'
        '<span class="rank-icon" aria-hidden="true">ðŸ†</span>'
        '<span>Ranking & Comparativo</span></button>'
    )
    ujs = json.dumps({u["key"]: {"sigla": u["sigla"], "nome": u["nome"], "ogu": u["ogu"], "fex": u["fex"],
                                 "logo": u["logo"], "accent": u["accent"]} for u in UNIDADES}, ensure_ascii=False)
    u0 = UNIDADES[0]
    posicao = periodo if periodo else (data_str[8:10] + "/" + data_str[5:7] + "/" + data_str[0:4])
    ger = datetime.datetime.now().strftime("%d/%m/%Y Ã s %H:%M")
    celdata_json = json.dumps(CEL, ensure_ascii=False).replace("</", "<\\/")
    ncdata_json = json.dumps(NCD, ensure_ascii=False).replace("</", "<\\/")
    daydata_json = json.dumps(DAY, ensure_ascii=False).replace("</", "<\\/")
    teladata_json = json.dumps(TELA, ensure_ascii=False).replace("</", "<\\/")
    return f"""<!doctype html><html lang="pt-BR" style="--accent:{u0['accent']}"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="color-scheme" content="light dark">
<meta name="description" content="Painel de CrÃ©dito DisponÃ­vel das OMDS da Base de Apoio LogÃ­stico do ExÃ©rcito â€” SIAFI / Tesouro Gerencial">
<title>CrÃ©dito DisponÃ­vel â€” OMDS Ba Ap Log Ex</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500;600;700&family=Newsreader:ital,opsz,wght@0,6..72,500;0,6..72,600;0,6..72,700;1,6..72,400&display=swap" rel="stylesheet">
<style>{CSS}</style></head>
<body>
<div class="bcms-bar" aria-hidden="true"></div>
<header class="topbar">
  <div class="brand">
    <img class="brasao" id="emblema" src="assets/logos/{u0['logo']}" alt="BrasÃ£o {esc(u0['sigla'])}" loading="eager">
    <div><h1 id="uTitulo">CrÃ©dito DisponÃ­vel â€” {esc(u0['sigla'])}</h1>
    <p class="subtitle"><span id="uNome">{esc(u0['nome'])}</span> Â· Tesouro Gerencial / SIAFI Â· <span id="uUasg">UASGs {u0['ogu']} (OGU) e {u0['fex']} (FEx)</span></p></div>
  </div>
  <div class="topbar-r">
    <div class="selo-wrap"><span class="selo"><span class="live-dot" aria-hidden="true"></span> PosiÃ§Ã£o {esc(posicao)}</span><span class="selo-delay">â± dados com ~24h de defasagem</span></div>
    <button class="theme" id="themeBtn" aria-pressed="false" aria-label="Alternar tema claro/escuro" onclick="bcmsTheme()" title="Alternar tema">
      <svg viewBox="0 0 24 24" class="ic-sun" aria-hidden="true"><circle cx="12" cy="12" r="4.5" style="fill:currentColor"/><g style="stroke:currentColor;stroke-width:1.8;stroke-linecap:round"><path d="M12 2v2.5M12 19.5v2.5M2 12h2.5M19.5 12h2.5M4.93 4.93l1.77 1.77M17.3 17.3l1.77 1.77M19.07 4.93l-1.77 1.77M6.7 17.3l-1.77 1.77"/></g></svg>
      <svg viewBox="0 0 24 24" class="ic-moon" aria-hidden="true"><path d="M20 14.5A8 8 0 019.5 4 8 8 0 1020 14.5z" style="fill:currentColor"/></svg>
    </button>
  </div>
</header>
<nav class="omds-nav" aria-label="Trocar de organizaÃ§Ã£o militar"><div class="omds-nav-in">{omds}</div></nav>
<main class="wrap">
  {banner}
  {"".join(frags)}
</main>
<div class="modal" id="modal" role="dialog" aria-modal="true" aria-labelledby="modal-title" aria-hidden="true" onclick="if(event.target===this)bcmsCelClose()">
  <div class="modal-panel">
    <button class="modal-x" aria-label="Fechar" onclick="bcmsCelClose()">âœ•</button>
    <div id="modal-body"></div>
  </div>
</div>
<footer class="rodape">
  <p class="rodape-brand">âš™ Comando da Base de Apoio LogÃ­stico do ExÃ©rcito Â· OMDS Subordinadas</p>
  <p><b>Metodologia:</b> CrÃ©dito DisponÃ­vel = ProvisÃ£o Recebida âˆ’ ProvisÃ£o Concedida âˆ’ Despesas Empenhadas (saldo lÃ­quido nÃ£o empenhado no Tesouro Gerencial / SIAFI). O detalhe Ã© o saldo real por cÃ©lula orÃ§amentÃ¡ria (AÃ§Ã£o Â· PI Â· ND). A soma das cÃ©lulas reconcilia com exatidÃ£o matemÃ¡tica com o total consolidado de cada OM.</p>
  <p>Fonte: CRÃ‰DITO DISP.xlsx (Google Drive / Tesouro Gerencial) Â· <b>â± Dados com defasagem de aproximadamente 24 horas.</b> Â· Painel atualizado em {esc(ger)}</p>
</footer>
<script>var CELDATA={celdata_json};var NCDATA={ncdata_json};var DAYDATA={daydata_json};var TELADATA={teladata_json};var UNIDADES={ujs};</script>
<script>{JS}</script>
</body></html>"""

# ============ CSS / JS (Constantes SÃªnior UI/UX) ============
CSS = r"""
/* ==========================================================================
   DESIGN TOKENS & MASTER UI/UX ARCHITECTURE (60fps GPU + WCAG 2.2 AAA)
   ========================================================================== */
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

:root {
  /* Escala Neutra HSL */
  --neutral-0:   #FFFFFF;
  --neutral-50:  #F8FAFC;
  --neutral-100: #F1F5F9;
  --neutral-200: #E2E8F0;
  --neutral-300: #CBD5E1;
  --neutral-400: #94A3B8;
  --neutral-500: #64748B;
  --neutral-600: #475569;
  --neutral-700: #334155;
  --neutral-800: #1E293B;
  --neutral-900: #0F172A;
  --neutral-950: #020617;

  /* SuperfÃ­cies & Fundo */
  --bg:          var(--neutral-50);
  --bg-surface:  var(--neutral-0);
  --bg-elevated: var(--neutral-0);
  --bg-subtle:   var(--neutral-100);

  /* Tipografia & Textos */
  --ink:         var(--neutral-900);
  --ink-muted:   var(--neutral-500);
  --ink-soft:    var(--neutral-400);

  /* Bordas */
  --border:        var(--neutral-200);
  --border-strong: var(--neutral-300);
  --border-focus:  #2563EB;

  /* Cores de Marca & PrimÃ¡rias */
  --primary:        #1C4A73;
  --primary-strong: #143A5C;
  --primary-50:     #EFF6FF;
  --primary-600:    #2563EB;

  /* Cores SemÃ¢nticas de Estado (WCAG AAA) */
  --success:        #059669;
  --success-strong: #047857;
  --success-bg:     #ECFDF5;
  --success-border: #A7F3D0;
  --hero-soft:      #F0FDF4;

  --warning:        #D97706;
  --warning-ink:    #92400E;
  --warning-main:   #D97706;
  --warning-bg:     #FFFBEB;
  --warning-border: #FDE68A;

  --danger:         #DC2626;
  --danger-main:    #DC2626;
  --danger-bg:      #FEF2F2;
  --danger-border:  #FECACA;

  --gold:           #D99B26;
  --gold-bg:        #FFFDF5;
  --gold-border:    #FDE68A;
  --gold-dark:      #92400E;

  --focus:          #059669;
  --track:          #E2E8F0;

  /* EstÃ¡gios Funil */
  --stg1: #1C4A73;
  --stg2: #3B82F6;
  --stg3: #10B981;

  /* Sombras FÃ­sicas em Camadas */
  --shadow-xs: 0 1px 2px rgba(15, 23, 42, 0.04);
  --shadow:    0 1px 3px rgba(15, 23, 42, 0.06), 0 1px 2px rgba(15, 23, 42, 0.04);
  --shadow-md: 0 4px 8px -2px rgba(15, 23, 42, 0.08), 0 2px 4px -2px rgba(15, 23, 42, 0.04);
  --shadow-h:  0 10px 20px -3px rgba(15, 23, 42, 0.10), 0 4px 6px -4px rgba(15, 23, 42, 0.05);
  --shadow-lg: 0 20px 25px -5px rgba(15, 23, 42, 0.10), 0 8px 10px -6px rgba(15, 23, 42, 0.04);

  /* Tipografia */
  --sans:  'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
  --mono:  'JetBrains Mono', Consolas, monospace;
  --serif: 'Newsreader', Georgia, serif;

  /* TransiÃ§Ãµes Spring */
  --ease-spring: cubic-bezier(0.16, 1, 0.3, 1);
  --ease-smooth: cubic-bezier(0.25, 1, 0.5, 1);
}

/* --- Dark Mode Elegante (OLED + Baixa Fadiga Ocular) --- */
@media (prefers-color-scheme: dark) {
  :root:not([data-theme="light"]) {
    --bg:          #090E17;
    --bg-surface:  #101926;
    --bg-elevated: #162234;
    --bg-subtle:   #1C2B40;

    --ink:         #F8FAFC;
    --ink-muted:   #94A3B8;
    --ink-soft:    #64748B;

    --border:        rgba(255, 255, 255, 0.08);
    --border-strong: rgba(255, 255, 255, 0.16);
    --border-focus:  #3B82F6;

    --primary:        #60A5FA;
    --primary-strong: #93C5FD;
    --primary-50:     #1E293B;
    --primary-600:    #3B82F6;

    --success:        #34D399;
    --success-strong: #34D399;
    --success-bg:     rgba(5, 150, 105, 0.15);
    --success-border: rgba(5, 150, 105, 0.35);
    --hero-soft:      rgba(5, 150, 105, 0.08);

    --warning:        #FBBF24;
    --warning-ink:    #FBBF24;
    --warning-main:   #F59E0B;
    --warning-bg:     rgba(217, 119, 6, 0.15);
    --warning-border: rgba(217, 119, 6, 0.35);

    --danger:         #F87171;
    --danger-main:    #EF4444;
    --danger-bg:      rgba(220, 38, 38, 0.15);
    --danger-border:  rgba(220, 38, 38, 0.35);

    --gold:           #F59E0B;
    --gold-bg:        rgba(217, 155, 38, 0.15);
    --gold-border:    rgba(217, 155, 38, 0.35);
    --gold-dark:      #FDE68A;

    --focus:          #34D399;
    --track:          #1E293B;

    --stg1: #60A5FA;
    --stg2: #93C5FD;
    --stg3: #34D399;

    --shadow-xs: none;
    --shadow:    0 2px 6px rgba(0, 0, 0, 0.4);
    --shadow-md: 0 4px 12px rgba(0, 0, 0, 0.5);
    --shadow-h:  0 10px 24px rgba(0, 0, 0, 0.6);
  }
}

:root[data-theme="dark"] {
  --bg:          #090E17;
  --bg-surface:  #101926;
  --bg-elevated: #162234;
  --bg-subtle:   #1C2B40;

  --ink:         #F8FAFC;
  --ink-muted:   #94A3B8;
  --ink-soft:    #64748B;

  --border:        rgba(255, 255, 255, 0.08);
  --border-strong: rgba(255, 255, 255, 0.16);
  --border-focus:  #3B82F6;

  --primary:        #60A5FA;
  --primary-strong: #93C5FD;
  --primary-50:     #1E293B;
  --primary-600:    #3B82F6;

  --success:        #34D399;
  --success-strong: #34D399;
  --success-bg:     rgba(5, 150, 105, 0.15);
  --success-border: rgba(5, 150, 105, 0.35);
  --hero-soft:      rgba(5, 150, 105, 0.08);

  --warning:        #FBBF24;
  --warning-ink:    #FBBF24;
  --warning-main:   #F59E0B;
  --warning-bg:     rgba(217, 119, 6, 0.15);
  --warning-border: rgba(217, 119, 6, 0.35);

  --danger:         #F87171;
  --danger-main:    #EF4444;
  --danger-bg:      rgba(220, 38, 38, 0.15);
  --danger-border:  rgba(220, 38, 38, 0.35);

  --gold:           #F59E0B;
  --gold-bg:        rgba(217, 155, 38, 0.15);
  --gold-border:    rgba(217, 155, 38, 0.35);
  --gold-dark:      #FDE68A;

  --focus:          #34D399;
  --track:          #1E293B;

  --stg1: #60A5FA;
  --stg2: #93C5FD;
  --stg3: #34D399;

  --shadow-xs: none;
  --shadow:    0 2px 6px rgba(0, 0, 0, 0.4);
  --shadow-md: 0 4px 12px rgba(0, 0, 0, 0.5);
  --shadow-h:  0 10px 24px rgba(0, 0, 0, 0.6);
}

:root[data-theme="light"] {
  --bg:          var(--neutral-50);
  --bg-surface:  var(--neutral-0);
  --bg-elevated: var(--neutral-0);
  --bg-subtle:   var(--neutral-100);
  --ink:         var(--neutral-900);
  --ink-muted:   var(--neutral-500);
  --border:        var(--neutral-200);
  --border-strong: var(--neutral-300);
  --hero-soft:      #F0FDF4;
}

html {
  transition: background-color .2s var(--ease-smooth), color .2s var(--ease-smooth);
  font-size: 16px;
  scroll-behavior: smooth;
}

body {
  background: var(--bg);
  color: var(--ink);
  font-family: var(--sans);
  font-size: 0.9375rem;
  line-height: 1.55;
  -webkit-font-smoothing: antialiased;
  -moz-osx-font-smoothing: grayscale;
}

.num {
  font-family: var(--mono);
  font-variant-numeric: tabular-nums lining-nums;
  font-feature-settings: "tnum" 1, "lnum" 1;
}

.visually-hidden {
  position: absolute;
  width: 1px; height: 1px;
  overflow: hidden;
  clip: rect(0 0 0 0);
}

.wrap {
  max-width: 1280px;
  margin: 0 auto;
  padding: 0 24px 64px;
}

.sec { margin-top: 36px; }
.eyebrow {
  font-size: 0.75rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.12em;
  color: var(--ink-muted);
  margin-bottom: 12px;
}

.sec-nota {
  font-size: 0.8125rem;
  color: var(--ink-muted);
  margin: -4px 0 16px;
  max-width: 960px;
  line-height: 1.6;
}
.sec-nota b { color: var(--ink); font-weight: 600; }

/* BotÃµes Excel Profissionais */
.btn-excel {
  position: relative;
  display: inline-flex;
  align-items: center;
  gap: 8px;
  background: linear-gradient(135deg, #107C41 0%, #0B5E31 100%);
  color: #FFFFFF !important;
  border: 1px solid #0E6B38;
  border-radius: 8px;
  padding: 0 16px;
  min-height: 48px;
  font-family: var(--sans);
  font-size: 0.8125rem;
  font-weight: 600;
  cursor: pointer;
  box-shadow: 0 2px 4px rgba(16, 124, 65, 0.2);
  transition: transform .2s var(--ease-spring);
  white-space: nowrap;
}
.btn-excel::after {
  content: '';
  position: absolute; inset: -1px;
  border-radius: inherit;
  background: linear-gradient(135deg, #148C4A 0%, #0E733D 100%);
  box-shadow: 0 6px 14px rgba(16, 124, 65, 0.32);
  opacity: 0;
  pointer-events: none;
  transition: opacity .2s var(--ease-smooth);
}
.btn-excel:hover {
  transform: translateY(-2px);
}
.btn-excel:hover::after {
  opacity: 1;
}
.btn-excel:active {
  transform: translateY(0) scale(0.97);
}
.btn-excel:active::after {
  opacity: 0;
}
.btn-excel-ic { font-size: 1rem; line-height: 1; }
.btn-excel-lg { padding: 10px 20px; font-size: 0.875rem; border-radius: 10px; }

/* Abas de Topo (Segmented Control Fluido) */
.toptabs {
  display: flex;
  gap: 6px;
  margin: 24px 0 8px;
  background: var(--bg-subtle);
  border: 1px solid var(--border);
  border-radius: 12px;
  padding: 5px;
}
.toptab {
  flex: 1;
  background: transparent;
  border: none;
  color: var(--ink-muted);
  font-family: var(--sans);
  font-size: 0.875rem;
  font-weight: 600;
  padding: 10px 16px;
  min-height: 48px;
  border-radius: 8px;
  cursor: pointer;
  transition: color .2s var(--ease-out-expo), background-color .2s var(--ease-out-expo), box-shadow .2s var(--ease-spring);
}
.toptab.on {
  background: var(--bg-surface);
  color: var(--ink);
  box-shadow: var(--shadow-md);
  font-weight: 700;
}
.toptab:hover:not(.on) { color: var(--ink); background: rgba(0,0,0,0.03); }
:root[data-theme="dark"] .toptab:hover:not(.on) { background: rgba(255,255,255,0.04); }

/* Topbar & Header */
.bcms-bar {
  height: 4px;
  background: linear-gradient(to right, var(--accent, #CE2B2B) 0%, var(--accent, #CE2B2B) 65%, var(--primary-600) 65%, var(--primary-600) 100%);
}
.topbar {
  position: sticky;
  top: 0;
  z-index: 40;
  background: color-mix(in srgb, var(--bg-surface) 88%, transparent);
  backdrop-filter: blur(14px);
  -webkit-backdrop-filter: blur(14px);
  border-bottom: 1px solid var(--border);
  padding: 16px 24px;
  display: flex;
  justify-content: space-between;
  align-items: center;
  gap: 16px;
}
.brand { display: flex; align-items: center; gap: 14px; }
.brasao {
  width: 48px;
  height: 48px;
  object-fit: contain;
  flex: none;
  filter: drop-shadow(0 2px 4px rgba(0, 0, 0, 0.12));
}
h1 {
  font-family: var(--serif);
  font-size: clamp(1.25rem, 1.1rem + 0.8vw, 1.75rem);
  font-weight: 700;
  letter-spacing: -0.01em;
  line-height: 1.15;
  color: var(--ink);
}
.subtitle {
  font-size: 0.78125rem;
  color: var(--ink-muted);
  margin-top: 3px;
}
.topbar-r { display: flex; align-items: center; gap: 12px; }
.selo-wrap { display: flex; flex-direction: column; align-items: flex-end; gap: 3px; }
.selo {
  display: inline-flex;
  align-items: center;
  gap: 6px;
  background: var(--bg-subtle);
  color: var(--ink-muted);
  border: 1px solid var(--border);
  border-radius: 999px;
  padding: 5px 12px;
  font-size: 0.78125rem;
  font-weight: 600;
  white-space: nowrap;
}
.live-dot {
  width: 7px;
  height: 7px;
  border-radius: 50%;
  background: var(--success);
  box-shadow: 0 0 8px var(--success);
}
.selo-delay { font-size: 0.6875rem; color: var(--warning-ink); font-weight: 600; white-space: nowrap; }

/* BotÃ£o Tema */
.theme {
  width: 48px;
  height: 48px;
  border: 1px solid var(--border);
  background: var(--bg-surface);
  border-radius: 9px;
  color: var(--ink-muted);
  cursor: pointer;
  display: grid;
  place-items: center;
  transition: border-color .2s var(--ease-out-expo), color .2s var(--ease-out-expo), transform .2s var(--ease-spring);
}
.theme:hover { border-color: var(--border-strong); color: var(--ink); transform: scale(1.05); }
.theme svg { width: 18px; height: 18px; }
.ic-moon { display: none; }
:root[data-theme="dark"] .ic-sun, html:not([data-theme]) .ic-sun { display: block; }
:root[data-theme="dark"] .ic-moon { display: block; }
:root[data-theme="dark"] .ic-sun { display: none; }
@media (prefers-color-scheme: dark) {
  html:not([data-theme]) .ic-sun { display: none; }
  html:not([data-theme]) .ic-moon { display: block; }
}

/* Barra de NavegaÃ§Ã£o OMDS */
.omds-nav {
  background: var(--bg-surface);
  border-bottom: 1px solid var(--border);
  position: relative;
  z-index: 30;
}
.omds-nav-in {
  max-width: 1280px;
  margin: 0 auto;
  padding: 10px 24px;
  display: flex;
  gap: 10px;
  overflow-x: auto;
  -webkit-overflow-scrolling: touch;
  scrollbar-width: thin;
}
.omds {
  flex: none;
  display: inline-flex;
  align-items: center;
  gap: 8px;
  padding: 7px 15px 7px 9px;
  min-height: 48px;
  border: 1px solid var(--border);
  border-radius: 999px;
  background: var(--bg-subtle);
  color: var(--ink-muted);
  font-family: var(--sans);
  font-size: 0.8125rem;
  font-weight: 700;
  cursor: pointer;
  white-space: nowrap;
  transition: transform .2s var(--ease-spring), border-color .2s var(--ease-out-expo), color .2s var(--ease-out-expo), background-color .2s var(--ease-out-expo);
}
.omds img { width: 22px; height: 22px; object-fit: contain; flex: none; }
.omds:hover { border-color: var(--border-strong); color: var(--ink); transform: translateY(-1px); }
.omds.on {
  color: #FFFFFF;
  background: var(--accent, #CE2B2B);
  border-color: var(--accent, #CE2B2B);
  box-shadow: 0 4px 12px color-mix(in srgb, var(--accent, #CE2B2B) 40%, transparent);
}
.omds.on img { filter: drop-shadow(0 0 2px rgba(255,255,255,0.7)); }

.omds-rank {
  background: linear-gradient(135deg, var(--bg-subtle) 0%, var(--gold-bg) 100%);
  border-color: var(--gold-border);
  color: var(--gold-dark);
}
.omds-rank.on {
  background: linear-gradient(135deg, #D99B26 0%, #B47810 100%) !important;
  border-color: #B47810 !important;
  color: #FFFFFF !important;
  box-shadow: 0 4px 14px rgba(217, 155, 38, 0.45);
}
.rank-icon { font-size: 0.9375rem; }

/* Cards & SuperfÃ­cies */
.card {
  position: relative;
  background: var(--bg-surface);
  border: 1px solid var(--border);
  border-radius: 14px;
  padding: 20px;
  box-shadow: var(--shadow);
  transition: border-color .2s var(--ease-out-expo);
}
.card:hover { border-color: var(--border-strong); }

/* Hero Card Moderno */
.hero {
  margin-top: 24px;
  background: var(--hero-soft);
  border: 1px solid var(--border);
  border-left: 5px solid var(--success);
  border-radius: 16px;
  padding: 24px 28px;
  display: grid;
  grid-template-columns: 1.35fr 1fr;
  gap: 28px;
  align-items: center;
  box-shadow: var(--shadow-md);
}
.hero-cmd {
  border-left-color: var(--gold);
  background: var(--gold-bg);
}
.hero-num {
  font-size: clamp(2.2rem, 1.8rem + 2vw, 3.25rem);
  font-weight: 800;
  letter-spacing: -0.025em;
  color: var(--success-strong);
  line-height: 1.05;
  margin: 6px 0 16px;
}
.hero-cmd .hero-num { color: var(--gold-dark); }

.hero-eq {
  display: flex;
  flex-wrap: wrap;
  align-items: center;
  gap: 8px;
}
.hero-eq-box {
  background: var(--bg-surface);
  border: 1px solid var(--border);
  border-radius: 8px;
  padding: 6px 12px;
  display: flex;
  flex-direction: column;
  gap: 2px;
  box-shadow: var(--shadow-xs);
}
.hero-eq-box .eq-tag {
  font-size: 0.625rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.08em;
  color: var(--ink-muted);
}
.hero-eq-box .eq-val {
  font-size: 0.9375rem;
  font-weight: 700;
  color: var(--ink);
}
.hero-eq-box.eq-highlight {
  border-color: var(--success-border);
  background: var(--success-bg);
}
.hero-eq-box.eq-highlight .eq-val { color: var(--success-strong); }
.hero-eq-sign {
  font-size: 1.25rem;
  font-weight: 300;
  color: var(--ink-muted);
  padding: 0 2px;
}

.delta {
  display: inline-flex;
  align-items: center;
  gap: 6px;
  font-size: 0.8125rem;
  font-weight: 600;
  margin-bottom: 12px;
  padding: 4px 12px;
  border-radius: 999px;
  background: var(--bg-subtle);
  border: 1px solid var(--border);
}
.delta.up { color: var(--success-strong); }
.delta.down { color: var(--danger); }
.delta.flat { color: var(--ink-muted); }
.delta small { font-weight: 400; color: var(--ink-muted); }

/* KPIs Grid */
.kpis {
  display: grid;
  grid-template-columns: repeat(4, 1fr);
  gap: 16px;
}
.kpi {
  position: relative;
  background: var(--bg-surface);
  border: 1px solid var(--border);
  border-left: 4px solid var(--border);
  border-radius: 12px;
  padding: 16px 18px;
  box-shadow: var(--shadow);
  display: flex;
  flex-direction: column;
  justify-content: space-between;
  transition: transform .2s var(--ease-spring);
}
.kpi::after {
  content: '';
  position: absolute;
  inset: -1px;
  border-radius: inherit;
  box-shadow: var(--shadow-md);
  opacity: 0;
  pointer-events: none;
  transition: opacity .2s var(--ease-out-expo);
}
.kpi:hover { transform: translateY(-2px); }
.kpi:hover::after { opacity: 1; }
.kpi-prov { border-left-color: var(--primary-600); }
.kpi-emp  { border-left-color: var(--warning-main); }
.kpi-liq  { border-left-color: var(--stg2); }
.kpi-pag  { border-left-color: var(--success); }
.kpi-l {
  font-size: 0.75rem;
  font-weight: 700;
  letter-spacing: 0.05em;
  color: var(--ink-muted);
  text-transform: uppercase;
}
.kpi-v {
  font-size: 1.5rem;
  font-weight: 700;
  margin: 6px 0;
  color: var(--ink);
}
.chip {
  display: inline-block;
  font-size: 0.6875rem;
  font-weight: 600;
  color: var(--ink-muted);
  background: var(--bg-subtle);
  border-radius: 999px;
  padding: 2px 10px;
  width: fit-content;
}

/* Grids */
.grid2 { display: grid; grid-template-columns: 1fr 1fr; gap: 18px; }

/* Cards de UASG */
.uasg { display: flex; flex-direction: column; gap: 10px; }
.uasg-h { display: flex; align-items: center; gap: 10px; flex-wrap: wrap; }
.uasg-cod { font-size: 1rem; font-weight: 700; }
.uasg-nome { font-size: 0.8125rem; color: var(--ink-muted); }
.uasg-disp { display: flex; justify-content: space-between; align-items: baseline; margin-top: 4px; }
.uasg-disp-l { font-size: 0.75rem; font-weight: 600; text-transform: uppercase; color: var(--ink-muted); }
.uasg-disp-v { font-size: 1.375rem; font-weight: 800; color: var(--success-strong); }
.uasg-eq { font-size: 0.78125rem; color: var(--ink-muted); }
.uasg-eq i { font-style: normal; color: var(--warning-ink); font-weight: 600; }
.uasg-exec { margin-top: 4px; }
.exec-l { display: flex; justify-content: space-between; font-size: 0.75rem; color: var(--ink-muted); margin-bottom: 5px; font-weight: 600; }
.exec-track { height: 8px; background: var(--track); border-radius: 999px; overflow: hidden; }
.exec-fill { height: 100%; background: var(--primary-600); border-radius: 999px; }

/* MÃ³dulo CrÃ©ditos em Tela por NC */
.et-head {
  display: flex;
  flex-wrap: wrap;
  gap: 14px;
  align-items: stretch;
  margin-bottom: 16px;
}
.et-kpi {
  background: var(--bg-surface);
  border: 1px solid var(--border);
  border-radius: 12px;
  padding: 14px 18px;
  display: flex;
  flex-direction: column;
  gap: 3px;
  min-width: 140px;
  box-shadow: var(--shadow);
}
.et-kpi span { font-size: 0.6875rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.06em; color: var(--ink-muted); }
.et-kpi b { font-size: 1.25rem; font-weight: 700; color: var(--ink); }
.et-hero { background: var(--hero-soft); border-left: 4px solid var(--success); }
.et-hero b { font-size: 1.5rem; color: var(--success-strong); }
.et-action { display: flex; align-items: center; }
.et-meta { margin-left: auto; align-self: center; font-size: 0.75rem; color: var(--ink-muted); text-align: right; line-height: 1.5; }

/* Badges de Idade de CrÃ©dito (SemÃ¡foro) */
.badge-age {
  display: inline-block;
  padding: 2px 8px;
  border-radius: 6px;
  font-size: 0.75rem;
  font-weight: 700;
  font-family: var(--mono);
}
.age-green { background: var(--success-bg); color: var(--success-strong); border: 1px solid var(--success-border); }
.age-amber { background: var(--warning-bg); color: var(--warning-ink); border: 1px solid var(--warning-border); }
.age-red   { background: var(--danger-bg);  color: var(--danger);       border: 1px solid var(--danger-border); }
.age-none  { color: var(--ink-muted); }

/* PÃ­lula de Fonte (160 / 167) */
.pill-fonte {
  display: inline-block;
  font-size: 0.6875rem;
  font-weight: 700;
  font-family: var(--mono);
  color: var(--primary);
  background: var(--bg-subtle);
  border: 1px solid var(--border-strong);
  border-radius: 5px;
  padding: 1px 7px;
}

/* Tabelas e Ferramentas */
.tabs {
  display: flex;
  gap: 6px;
  border-bottom: 1px solid var(--border);
  margin-bottom: 16px;
  overflow-x: auto;
}
.tab {
  background: none;
  border: none;
  border-bottom: 2px solid transparent;
  color: var(--ink-muted);
  font-family: var(--sans);
  font-size: 0.8125rem;
  font-weight: 600;
  padding: 10px 16px;
  cursor: pointer;
  white-space: nowrap;
  transition: transform .2s var(--ease-out-expo), color .2s var(--ease-out-expo), background-color .2s var(--ease-out-expo), border-color .2s var(--ease-out-expo);
}
.tab.on { color: var(--success-strong); border-bottom-color: var(--success); font-weight: 700; }
.tab:hover:not(.on) { color: var(--ink); }

.tbl-tools {
  display: flex;
  align-items: center;
  gap: 12px;
  margin-bottom: 12px;
  flex-wrap: wrap;
}
.tbl-search {
  flex: 1;
  min-width: 240px;
  background: var(--bg-surface);
  border: 1px solid var(--border);
  border-radius: 10px;
  padding: 10px 14px;
  color: var(--ink);
  font-family: var(--sans);
  font-size: 0.875rem;
  box-shadow: var(--shadow-xs);
  transition: border-color .2s var(--ease-out-expo), box-shadow .2s var(--ease-out-expo);
}
.tbl-search:focus {
  outline: 2px solid var(--border-focus);
  outline-offset: 1px;
  border-color: var(--border-focus);
}
.tbl-count { font-size: 0.78125rem; color: var(--ink-muted); margin-left: auto; font-weight: 500; }
.tbl-scroll {
  overflow-x: auto;
  border: 1px solid var(--border);
  border-radius: 14px;
  background: var(--bg-surface);
  box-shadow: var(--shadow);
}

table.det { border-collapse: collapse; width: 100%; font-size: 0.875rem; }
.det th {
  position: sticky;
  top: 0;
  z-index: 10;
  background: var(--bg-subtle);
  color: var(--ink-muted);
  font-size: 0.6875rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.06em;
  text-align: left;
  padding: 12px 14px;
  cursor: pointer;
  white-space: nowrap;
  border-bottom: 1px solid var(--border);
  user-select: none;
}
.det th.num, .det td.num { text-align: right; }
.det th .sort { display: inline-block; width: 12px; color: var(--success); }
.det td {
  padding: 10px 14px;
  border-bottom: 1px solid var(--border);
  color: var(--ink);
}
.det td.num { font-weight: 500; }
.det td.anchor { font-weight: 700; }
.det tbody tr:hover { background: var(--bg-subtle); }
.det .obj { max-width: 360px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; color: var(--ink-muted); }
/* [COMPACTA] lista de NC: linhas mais baixas e mais texto Ãºtil visÃ­vel.
   Reduz padding/fonte (46px -> ~32px por linha) e devolve Ã  descriÃ§Ã£o o espaÃ§o
   liberado pelo nÂº curto da NC. */
.det-compact td { padding: 5px 10px; line-height: 1.35; }
.det-compact th { padding: 8px 10px; }
.det-compact tfoot td { padding: 8px 10px; }
.det-compact .obj { max-width: 560px; color: var(--ink); }
.det-compact .mono2 { font-size: 0.78125rem; }
.det-compact .badge-age { padding: 1px 7px; font-size: 0.6875rem; }
.det-compact .pill-fonte { padding: 1px 7px; font-size: 0.6875rem; }
.nc-num { font-weight: 700; letter-spacing: .01em; }
.nc-ug { color: var(--ink-muted); font-weight: 500; }
/* selo de crÃ©dito vindo de mudanÃ§a de ND (objeto herdado da NC de origem) */
.tag-nd {
  display: inline-block; padding: 0 6px; margin-right: 5px; border-radius: 5px;
  background: var(--warning-bg, rgba(181,130,43,.14)); color: var(--warning-ink, #8A631C);
  font-size: 0.6875rem; font-weight: 700; white-space: nowrap; vertical-align: 1px;
}
/* ---- barra de filtros da lista em tela ---- */
.tbl-filtros {
  display: flex; flex-wrap: wrap; align-items: center; gap: 8px;
  padding: 10px 12px; margin-bottom: 10px; border-radius: 10px;
  background: var(--surface-2); border: 1px solid var(--border);
}
.flt-lbl { font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: .06em; color: var(--ink-muted); }
.flt {
  font-family: inherit; font-size: 0.8125rem; color: var(--ink);
  background: var(--surface); border: 1px solid var(--border-strong);
  border-radius: 8px; padding: 6px 10px; max-width: 240px; cursor: pointer;
}
.flt:focus-visible { outline: 2px solid var(--focus); outline-offset: 1px; }
.flt-limpa {
  font-family: inherit; font-size: 0.75rem; font-weight: 600; color: var(--ink-muted);
  background: transparent; border: 1px solid var(--border-strong); border-radius: 8px;
  padding: 6px 10px; cursor: pointer;
}
.flt-limpa:hover { color: var(--danger); border-color: var(--danger); }
.flt-resumo { font-size: 0.8125rem; color: var(--ink-muted); font-weight: 600; }
.flt-resumo.on { color: var(--primary); }
@media (max-width: 640px) { .flt { flex: 1 1 100%; max-width: none; } }
/* ---- modal por etapas do crÃ©dito ---- */
.m-etapas { display: flex; flex-wrap: wrap; gap: 6px; margin: 14px 0 4px; border-bottom: 1px solid var(--border); padding-bottom: 10px; }
.m-etapa {
  border: 1px solid var(--border); background: var(--bg-subtle); color: var(--ink-muted);
  border-radius: 999px; padding: 6px 14px; font-size: 0.8125rem; font-weight: 600;
  cursor: pointer; transition: transform .2s var(--ease-out-expo), color .2s var(--ease-out-expo), background-color .2s var(--ease-out-expo), border-color .2s var(--ease-out-expo); font-family: inherit;
}
.m-etapa:hover { color: var(--ink); border-color: var(--border-strong); }
.m-etapa.on { background: var(--primary); border-color: var(--primary); color: #fff; }
.m-etapa-body { padding-top: 14px; }
.m-tela-card {
  display: flex; flex-direction: column; gap: 4px; padding: 18px 20px; border-radius: 14px;
  background: var(--hero-soft, rgba(15,122,90,.10)); border: 1px solid var(--success, #0F7A5A);
}
.m-tela-lbl { font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: .06em; color: var(--success-strong, #0B5E45); }
.m-tela-val { font-size: 2rem; font-weight: 800; line-height: 1.1; font-variant-numeric: tabular-nums; color: var(--ink); }
.m-tela-meta { font-size: 0.8125rem; color: var(--ink-muted); }
.m-org { margin-top: 14px; border: 1px dashed var(--warning, #B5822B); border-radius: 12px; overflow: hidden; }
.m-org-h { padding: 8px 14px; background: var(--warning-bg, rgba(181,130,43,.12)); font-size: 0.8125rem; color: var(--warning-ink, #8A631C); }
.m-org-b { padding: 10px 14px; }
.m-org-b span { font-size: 0.75rem; color: var(--ink-muted); }
.m-org-b p { margin-top: 4px; font-size: 0.875rem; color: var(--ink); }
.m-nc-desc.big { font-size: 0.9375rem; line-height: 1.5; color: var(--ink); }
.m-nc-esta { border-color: var(--primary); box-shadow: 0 0 0 1px var(--primary) inset; }
.m-nc-tag { font-style: normal; font-size: 0.625rem; font-weight: 700; text-transform: uppercase;
  background: var(--primary); color: #fff; padding: 1px 6px; border-radius: 4px; margin-left: 6px; }
.m-barra { margin: 14px 0 4px; }
.m-barra-t { display: flex; justify-content: space-between; font-size: 0.8125rem; margin-bottom: 5px; color: var(--ink-muted); }
.m-barra-t b { color: var(--ink); font-variant-numeric: tabular-nums; }
.m-barra-track { height: 10px; border-radius: 999px; background: var(--track, #E4E8EF); overflow: hidden; }
.m-barra-fill { height: 100%; border-radius: 999px; background: var(--success, #0F7A5A); }
.det .mono2 { font-size: 0.8125rem; color: var(--ink-muted); }
.cell-neg { color: var(--danger); background: var(--danger-bg); }
.det tfoot td { padding: 12px 14px; font-weight: 700; background: var(--bg-subtle); border-top: 2px solid var(--border-strong); }
.cel-row { cursor: pointer; }
.cel-row .chev { float: right; margin-left: 8px; color: var(--ink-soft); font-weight: 400; transition: transform .15s ease, color .15s ease; }
.cel-row:hover .chev { color: var(--success); transform: translateX(3px); }

/* PÃ³dio & Benchmarking (Ranking OMDS) */
.ranking-header-card {
  margin-top: 24px;
  padding: 26px 30px;
  background: var(--bg-surface);
  border: 1px solid var(--border);
  border-top: 5px solid var(--gold);
  border-radius: 16px;
  box-shadow: var(--shadow-md);
}
.rh-tag { font-size: 0.6875rem; font-weight: 800; letter-spacing: 0.14em; color: var(--gold-dark); margin-bottom: 6px; }
.rh-title { font-family: var(--serif); font-size: clamp(1.4rem, 1.2rem + 1vw, 2rem); font-weight: 700; margin-bottom: 8px; line-height: 1.2; }
.rh-desc { font-size: 0.875rem; color: var(--ink-muted); max-width: 900px; line-height: 1.6; }

.podium-wrap {
  display: grid;
  grid-template-columns: repeat(3, 1fr);
  gap: 20px;
  margin-top: 18px;
  align-items: end;
}
.podium-step {
  position: relative;
  background: var(--bg-surface);
  border: 1px solid var(--border);
  border-radius: 16px;
  padding: 24px 20px 20px;
  text-align: center;
  cursor: pointer;
  box-shadow: var(--shadow);
  transition: opacity .25s var(--ease-out-expo), transform .25s var(--ease-spring);
  display: flex;
  flex-direction: column;
  align-items: center;
}
.podium-step:hover {
  transform: translateY(-6px);
  box-shadow: var(--shadow-h);
  border-color: var(--border-strong);
}
.podium-gold {
  border-top: 6px solid #F59E0B;
  background: linear-gradient(to bottom, var(--gold-bg), var(--bg-surface));
  order: 2;
  padding-top: 32px;
  margin-bottom: 16px;
}
.podium-silver {
  border-top: 6px solid #94A3B8;
  background: linear-gradient(to bottom, var(--bg-subtle), var(--bg-surface));
  order: 1;
}
.podium-bronze {
  border-top: 6px solid #CD7F32;
  background: linear-gradient(to bottom, color-mix(in srgb, #CD7F32 8%, var(--bg-surface)), var(--bg-surface));
  order: 3;
}
.podium-badge {
  display: inline-block;
  font-size: 0.75rem;
  font-weight: 800;
  padding: 4px 14px;
  border-radius: 999px;
  background: var(--bg-subtle);
  border: 1px solid var(--border);
  margin-bottom: 14px;
}
.podium-gold .podium-badge { background: #FFF5D6; color: #855A00; border-color: #F0D070; }
:root[data-theme="dark"] .podium-gold .podium-badge { background: #3B2C08; color: #F0D070; border-color: #6E5110; }

.podium-avatar-wrap {
  width: 68px;
  height: 68px;
  border-radius: 50%;
  background: var(--bg-surface);
  border: 2px solid var(--border);
  display: flex;
  align-items: center;
  justify-content: center;
  margin-bottom: 12px;
  overflow: hidden;
  padding: 6px;
  box-shadow: var(--shadow-xs);
}
.podium-gold .podium-avatar-wrap {
  width: 80px;
  height: 80px;
  border-color: #F59E0B;
  box-shadow: 0 0 20px rgba(245, 158, 11, 0.35);
}
.podium-logo { width: 100%; height: 100%; object-fit: contain; }
.podium-sigla { font-size: 1.1875rem; font-weight: 800; letter-spacing: -0.01em; margin-bottom: 2px; }
.podium-nome { font-size: 0.75rem; color: var(--ink-muted); margin-bottom: 16px; max-width: 210px; line-height: 1.35; height: 32px; overflow: hidden; }
.podium-stat-pill {
  background: var(--bg-subtle);
  border: 1px solid var(--border);
  border-radius: 10px;
  padding: 10px 16px;
  width: 100%;
  display: flex;
  justify-content: space-between;
  align-items: center;
  margin-bottom: 10px;
}
.podium-stat-pill .stat-l { font-size: 0.6875rem; text-transform: uppercase; color: var(--ink-muted); font-weight: 700; }
.podium-stat-pill .stat-v { font-size: 1.125rem; font-weight: 800; color: var(--success-strong); }
.podium-substat { font-size: 0.75rem; color: var(--ink-muted); margin-bottom: 16px; }
.podium-btn {
  background: var(--bg-subtle);
  border: 1px solid var(--border);
  border-radius: 8px;
  padding: 8px 14px;
  font-family: var(--sans);
  font-size: 0.75rem;
  font-weight: 700;
  color: var(--primary);
  cursor: pointer;
  transition: transform .2s var(--ease-out-expo), background-color .2s var(--ease-out-expo);
  width: 100%;
}
.podium-btn:hover { background: var(--primary); color: #FFFFFF; border-color: var(--primary); }

/* Tabela Comparativa */
.tbl-om-cell { display: flex; align-items: center; gap: 10px; }
.tbl-om-logo { width: 28px; height: 28px; object-fit: contain; flex: none; }
.tbl-om-sub { font-size: 0.75rem; color: var(--ink-muted); font-weight: 400; display: block; }
.tbl-pct-cell { display: flex; align-items: center; gap: 8px; justify-content: flex-end; }
.mini-track { width: 64px; height: 6px; background: var(--track); border-radius: 3px; overflow: hidden; }
.mini-fill { height: 100%; border-radius: 3px; }
.tbl-action-btn {
  background: var(--bg-subtle);
  border: 1px solid var(--border);
  border-radius: 6px;
  padding: 5px 10px;
  font-size: 0.75rem;
  font-weight: 700;
  color: var(--primary);
  cursor: pointer;
}
.tbl-action-btn:hover { background: var(--primary); color: #FFFFFF; }

/* SVG Classes */
.svg { width: 100%; height: auto; display: block; }
.s-lbl { font-size: 11px; font-weight: 700; letter-spacing: 0.06em; fill: var(--ink-muted); }
.s-seg { font-size: 11px; fill: var(--warning-ink); font-weight: 600; font-family: var(--mono); }
.s-seg-ok { fill: var(--success-strong); }
.s-brk { stroke: var(--border-strong); stroke-width: 1; }
.s-cat { font-size: 12px; fill: var(--ink-muted); }
.s-cat-ok { fill: var(--success-strong); font-weight: 600; }
.s-val { font-size: 13px; font-weight: 700; font-family: var(--mono); }
.s-on { fill: #FFFFFF; }
.s-ok { fill: var(--success-strong); }
.s-warn { fill: var(--warning-ink); }
.s-on2 { fill: var(--ink); }
.s-num { font-size: 11px; font-weight: 600; font-family: var(--mono); }
.s-neg { fill: var(--danger); }
.s-conn { stroke: var(--border-strong); stroke-width: 1; stroke-dasharray: 4 3; }
.s-zero { stroke: var(--border-strong); stroke-width: 1.5; }
.s-grid { stroke: var(--border); stroke-width: 1; }
.s-ax { font-size: 11px; fill: var(--ink-muted); font-family: var(--mono); }
.s-line { fill: none; stroke: var(--success); stroke-width: 2.5; }
.s-area { fill: var(--success); opacity: .12; }
.s-dot { fill: var(--success); }
.s-conv { font-size: 10.5px; fill: var(--ink-muted); }

/* Modal Drill-Down (FÃ­sica Spring + Backdrop Blur) */
.modal {
  position: fixed;
  inset: 0;
  z-index: 999;
  display: none;
  align-items: center;
  justify-content: center;
  padding: 24px 16px;
  background: rgba(2, 6, 23, 0.65);
  backdrop-filter: blur(10px);
  -webkit-backdrop-filter: blur(10px);
  overflow-y: auto;
}
.modal.open { display: flex; }
.modal-panel {
  position: relative;
  background: var(--bg-surface);
  border: 1px solid var(--border-strong);
  border-radius: 18px;
  box-shadow: var(--shadow-lg);
  max-width: 780px;
  width: 100%;
  padding: 28px 30px;
  animation: modalIn .25s var(--ease-spring);
  /* [FIX janela fora da tela] o painel NUNCA excede a viewport: limita a altura e
     rola o conteÃºdo por dentro (o âœ• fica sempre visÃ­vel). Antes crescia sem limite
     e, com align-items:center, o topo era cortado e ficava inalcanÃ§Ã¡vel. */
  max-height: calc(100vh - 48px);
  max-height: calc(100dvh - 48px);
  margin: auto;
  display: flex;
  flex-direction: column;
  min-height: 0;
}
#modal-body {
  overflow-y: auto;
  overscroll-behavior: contain;
  min-height: 0;
  margin-right: -10px;
  padding-right: 10px;
}
@keyframes modalIn {
  from { opacity: 0; transform: scale(0.96) translateY(10px); }
  to { opacity: 1; transform: scale(1) translateY(0); }
}
.modal-x {
  position: absolute;
  top: 16px;
  right: 16px;
  width: 36px;
  height: 36px;
  border: 1px solid var(--border);
  background: var(--bg-subtle);
  border-radius: 10px;
  color: var(--ink-muted);
  cursor: pointer;
  font-size: 16px;
  line-height: 1;
  display: grid;
  place-items: center;
  transition: transform .2s var(--ease-out-expo), color .2s var(--ease-out-expo), background-color .2s var(--ease-out-expo), border-color .2s var(--ease-out-expo);
}
.modal-x:hover { color: var(--ink); border-color: var(--border-strong); transform: scale(1.05); }
#modal-body h3 { font-size: 1.25rem; font-weight: 700; margin-bottom: 4px; padding-right: 44px; }
.m-sub { font-size: 0.8125rem; color: var(--ink-muted); margin-bottom: 16px; }
.m-ficha {
  display: flex;
  flex-wrap: wrap;
  gap: 10px 20px;
  margin: 4px 0 16px;
  padding: 14px 16px;
  background: var(--bg-subtle);
  border: 1px solid var(--border);
  border-radius: 12px;
}
.m-ficha span { display: flex; flex-direction: column; gap: 2px; font-size: 0.6875rem; text-transform: uppercase; letter-spacing: 0.05em; color: var(--ink-muted); min-width: 120px; }
.m-ficha span.wide { flex-basis: 100%; }
.m-ficha b { font-size: 0.875rem; font-weight: 600; color: var(--ink); text-transform: none; }
.m-kpis { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 16px; }
.m-kpis span {
  flex: 1;
  min-width: 120px;
  display: flex;
  flex-direction: column;
  gap: 2px;
  font-size: 0.6875rem;
  color: var(--ink-muted);
  text-transform: uppercase;
  letter-spacing: 0.04em;
  background: var(--bg-subtle);
  border: 1px solid var(--border);
  border-radius: 10px;
  padding: 10px 12px;
}
.m-kpis b { font-size: 1.125rem; font-weight: 700; color: var(--ink); font-family: var(--mono); }
.m-kpis .ok b { color: var(--success-strong); }
.m-kpis .ok { border-left: 3px solid var(--success); }
.m-formula { font-size: 0.75rem; color: var(--ink-muted); margin: -8px 0 16px; font-style: italic; }
.m-ncs-h { font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; color: var(--ink-muted); margin-bottom: 10px; }
.m-ncs { display: flex; flex-direction: column; gap: 10px; max-height: 46vh; overflow-y: auto; }
.m-nc { border: 1px solid var(--border); border-radius: 12px; padding: 12px 14px; background: var(--bg-surface); }
.m-nc-h { display: flex; justify-content: space-between; align-items: baseline; gap: 12px; }
.m-nc-num { font-weight: 700; font-size: 0.875rem; font-family: var(--mono); }
.m-nc-val { font-weight: 700; font-size: 0.875rem; color: var(--success-strong); white-space: nowrap; font-family: var(--mono); }
.m-nc-val.neg { color: var(--danger); }
.m-nc-op { font-size: 0.6875rem; text-transform: uppercase; letter-spacing: 0.04em; color: var(--ink-muted); margin-top: 4px; }
.m-nc-desc { font-size: 0.8125rem; color: var(--ink); margin-top: 6px; line-height: 1.55; white-space: pre-wrap; word-break: break-word; }

/* Toast Feedback */
.toast {
  position: fixed;
  bottom: 28px;
  right: 28px;
  background: var(--neutral-900);
  color: #FFFFFF;
  padding: 14px 22px;
  border-radius: 12px;
  box-shadow: var(--shadow-lg);
  font-size: 0.875rem;
  font-weight: 600;
  display: flex;
  align-items: center;
  gap: 12px;
  z-index: 9999;
  opacity: 0;
  transform: translateY(16px);
  transition: opacity .25s var(--ease-out-expo), transform .25s var(--ease-spring);
  pointer-events: none;
  border: 1px solid rgba(255, 255, 255, 0.15);
}
:root[data-theme="dark"] .toast { background: var(--bg-elevated); color: var(--ink); border-color: var(--border-strong); }
.toast.show { opacity: 1; transform: translateY(0); }
.toast-ic { font-size: 1.125rem; }

/* Banner de Alerta */
.banner {
  background: var(--danger-bg);
  border: 1px solid var(--danger-border);
  border-radius: 12px;
  padding: 14px 18px;
  margin-top: 20px;
  font-size: 0.8125rem;
  color: var(--danger);
}
.banner ul { margin: 6px 0 0 20px; }

/* RodapÃ© */
.rodape {
  max-width: 1280px;
  margin: 48px auto 0;
  padding: 24px;
  border-top: 1px solid var(--border);
  color: var(--ink-muted);
  font-size: 0.75rem;
  line-height: 1.7;
}
.rodape b { color: var(--ink); }
.rodape-brand { color: var(--gold-dark); font-weight: 700; font-size: 0.8125rem; letter-spacing: 0.02em; margin-bottom: 8px; }

/* Acessibilidade & Estados de Foco */
:focus-visible { outline: 2px solid var(--border-focus); outline-offset: 2px; border-radius: 6px; }
@media (prefers-reduced-motion: reduce) {
  *, *::before, *::after { transition: none !important; animation: none !important; }
}

/* Responsividade Mobile-First */
@media (max-width: 1023px) {
  .grid2 { grid-template-columns: 1fr; }
  .hero { grid-template-columns: 1fr; }
  .podium-wrap { grid-template-columns: 1fr; gap: 14px; }
  .podium-gold { order: 1; margin-bottom: 0; padding-top: 24px; }
  .podium-silver { order: 2; }
  .podium-bronze { order: 3; }
  .kpis { grid-template-columns: repeat(2, 1fr); }
}

@media (max-width: 640px) {
  .wrap { padding: 0 16px 48px; }
  .topbar { padding: 12px 16px; flex-wrap: wrap; }
  h1 { font-size: 1.25rem; }
  .subtitle { display: none; }
  .hero-num { font-size: 2.2rem; }
  .kpis { grid-template-columns: 1fr; }
  .det td.mono2, .det th:first-child, .det td:first-child { position: sticky; left: 0; background: var(--bg-surface); }
  .det th:first-child { background: var(--bg-subtle); }
}
"""

JS = r"""
(function(){
  var s=localStorage.getItem('bcms-theme');
  if(s){ document.documentElement.setAttribute('data-theme',s); }
})();

function bcmsTheme(){
  var h=document.documentElement;
  var cur=h.getAttribute('data-theme');
  var dark=cur?cur==='dark':window.matchMedia('(prefers-color-scheme:dark)').matches;
  var next=dark?'light':'dark';
  h.setAttribute('data-theme',next);
  localStorage.setItem('bcms-theme',next);
  var btn=document.getElementById('themeBtn');
  if(btn) btn.setAttribute('aria-pressed',next==='dark');
}

function bcmsTab(btn,id){
  var list=btn.parentNode.querySelectorAll('.tab');
  list.forEach(function(b){b.classList.remove('on');b.setAttribute('aria-selected','false');b.tabIndex=-1;});
  btn.classList.add('on');btn.setAttribute('aria-selected','true');btn.tabIndex=0;
  (btn.closest('.unidade')||document).querySelectorAll('.tabpanel').forEach(function(p){p.style.display='none';});
  var target=document.getElementById(id);
  if(target) target.style.display='block';
}

function bcmsTabKey(e,btn){
  var t=Array.prototype.slice.call(btn.parentNode.querySelectorAll('.tab'));
  var i=t.indexOf(btn);
  if(e.key==='ArrowRight'){e.preventDefault();t[(i+1)%t.length].focus();t[(i+1)%t.length].click();}
  else if(e.key==='ArrowLeft'){e.preventDefault();t[(i-1+t.length)%t.length].focus();t[(i-1+t.length)%t.length].click();}
}

function bcmsView(btn,which){
  var m=btn.closest('.unidade')||document;
  btn.parentNode.querySelectorAll('.toptab').forEach(function(b){b.classList.remove('on');b.setAttribute('aria-selected','false');});
  btn.classList.add('on');btn.setAttribute('aria-selected','true');
  var vr=m.querySelector('.view-resumo');if(vr)vr.style.display=which==='resumo'?'':'none';
  var vc=m.querySelector('.view-completo');if(vc)vc.style.display=which==='completo'?'':'none';
  window.scrollTo({top:0,behavior:'smooth'});
}

function trocaOMDS(btn){
  var key=btn.getAttribute('data-key');
  document.querySelectorAll('.omds-nav .omds').forEach(function(b){b.classList.remove('on');b.setAttribute('aria-current','false');});
  btn.classList.add('on');btn.setAttribute('aria-current','true');
  document.querySelectorAll('.unidade').forEach(function(s){s.style.display=(s.getAttribute('data-key')===key)?'':'none';});
  if(key==='RANKING'){
    document.documentElement.style.setProperty('--accent','#D99B26');
    var em=document.getElementById('emblema');if(em){em.src='assets/logos/BAAPLOG.png';em.alt='BrasÃ£o Base de Apoio LogÃ­stico';}
    var t=document.getElementById('uTitulo');if(t)t.textContent='CrÃ©dito DisponÃ­vel â€” Ranking & Comparativo OMDS';
    var n=document.getElementById('uNome');if(n)n.textContent='Base de Apoio LogÃ­stico do ExÃ©rcito';
    var uu=document.getElementById('uUasg');if(uu)uu.textContent='Consolidado das 6 OrganizaÃ§Ãµes Militares Diretamente Subordinadas';
    try{document.title='CrÃ©dito DisponÃ­vel â€” Ranking & Comparativo OMDS';}catch(e){}
  } else {
    var u=UNIDADES[key];if(!u)return;
    document.documentElement.style.setProperty('--accent',u.accent);
    var em=document.getElementById('emblema');if(em){em.src='assets/logos/'+u.logo;em.alt='BrasÃ£o '+u.sigla;}
    var t=document.getElementById('uTitulo');if(t)t.textContent='CrÃ©dito DisponÃ­vel â€” '+u.sigla;
    var n=document.getElementById('uNome');if(n)n.textContent=u.nome;
    var uu=document.getElementById('uUasg');if(uu)uu.textContent='UASGs '+u.ogu+' (OGU) e '+u.fex+' (FEx)';
    try{document.title='CrÃ©dito DisponÃ­vel â€” '+u.sigla;}catch(e){}
  }
  try{localStorage.setItem('bcms-omds',key);}catch(e){}
  window.scrollTo({top:0,behavior:'smooth'});
}

function trocaOMDSPorKey(key){
  var btn=document.querySelector('.omds-nav .omds[data-key="'+key+'"]');
  if(btn){btn.click();window.scrollTo({top:0,behavior:'smooth'});}
}

(function(){
  try{
    var k=localStorage.getItem('bcms-omds');
    if(k){
      var b=document.querySelector('.omds-nav .omds[data-key="'+k+'"]');
      if(b&&!b.classList.contains('on')) b.click();
    }
  }catch(e){}
})();

function bcmsSearch(inp,tid){
  var q=inp.value.toLowerCase();
  var container=document.getElementById(tid);
  if(!container)return;
  var tb=container.querySelector('tbody');
  if(!tb)return;
  var rows=tb.querySelectorAll('tr');
  var n=0;
  rows.forEach(function(r){
    var ok=r.textContent.toLowerCase().indexOf(q)>-1;
    r.style.display=ok?'':'none';
    if(ok) n++;
  });
  var cnt=document.getElementById('cnt-'+tid);
  if(cnt){
    var u=cnt.getAttribute('data-unit')||'linha(s)';
    cnt.textContent=n+' '+u;
  }
}

function bcmsNC(row){
  var d=NCDATA[row.getAttribute('data-nc')];
  if(!d)return;
  var neg=d.val<0;
  var h='<h3 id="modal-title">NC '+bcmsEsc(d.nc)+'</h3>';
  h+='<p class="m-sub">'+bcmsEsc((d.u?d.u+' Â· ':'')+'AÃ§Ã£o '+d.acao+' Â· PI '+d.pi+' Â· ND '+d.nd+(d.ndn?' â€” '+d.ndn:''))+'</p>';
  h+='<div class="m-kpis"><span>Data<b>'+bcmsEsc(d.dia||'â€”')+'</b></span><span>OperaÃ§Ã£o<b class="op">'+bcmsEsc(d.op||'â€”')+'</b></span><span class="'+(neg?'':'ok')+'">Valor<b class="'+(neg?'neg':'')+'">'+bcmsBRL(d.val)+'</b></span></div>';
  h+='<div class="m-ncs-h">DescriÃ§Ã£o completa do lanÃ§amento</div>';
  h+='<div class="m-nc"><div class="m-nc-desc">'+bcmsEsc(d.obj||'(sem descriÃ§Ã£o)')+'</div></div>';
  document.getElementById('modal-body').innerHTML=h;
  var m=document.getElementById('modal');
  m.classList.add('open');
  m.setAttribute('aria-hidden','false');
  var x=document.querySelector('.modal-x');
  if(x) x.focus();
}

function bcmsSort(th){
  var table=th.closest('table');
  var idx=Array.prototype.indexOf.call(th.parentNode.children,th);
  var dir=th.getAttribute('aria-sort')==='ascending'?'descending':'ascending';
  th.parentNode.querySelectorAll('th').forEach(function(h){h.setAttribute('aria-sort','none');h.querySelector('.sort').textContent='';});
  th.setAttribute('aria-sort',dir);
  th.querySelector('.sort').textContent=dir==='ascending'?' â–²':' â–¼';
  var tb=table.querySelector('tbody');
  var rows=Array.prototype.slice.call(tb.querySelectorAll('tr'));
  rows.sort(function(a,b){
    var ca=a.children[idx], cb=b.children[idx];
    var da=ca.getAttribute('data-sort'), db=cb.getAttribute('data-sort');
    var va,vb;
    if(da!==null&&db!==null){va=parseFloat(da);vb=parseFloat(db);}
    else{va=ca.textContent.trim().toLowerCase();vb=cb.textContent.trim().toLowerCase();}
    if(va<vb)return dir==='ascending'?-1:1;
    if(va>vb)return dir==='ascending'?1:-1;
    return 0;
  });
  rows.forEach(function(r){tb.appendChild(r);});
}

function bcmsEsc(s){return String(s==null?'':s).replace(/[&<>"]/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c];});}
function bcmsEscXml(s){return String(s==null?'':s).replace(/[&<>"']/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&apos;'}[c];});}
function bcmsBRL(v){var neg=v<0,s=Math.abs(v).toFixed(2).split('.');var i=s[0].replace(/\B(?=(\d{3})+(?!\d))/g,'.');return (neg?'âˆ’R$ ':'R$ ')+i+','+s[1];}

function bcmsToast(msg){
  var old=document.getElementById('bcms-toast');if(old)old.remove();
  var t=document.createElement('div');t.id='bcms-toast';t.className='toast';
  t.innerHTML='<span class="toast-ic">âœ¨</span><span>'+bcmsEsc(msg)+'</span>';
  document.body.appendChild(t);
  setTimeout(function(){t.classList.add('show');},10);
  setTimeout(function(){t.classList.remove('show');setTimeout(function(){if(t.parentNode)t.remove();},300);},3500);
}

/* ---- Filtros combinados da lista "em tela" (Fonte + AÃ§Ã£o + ND + Idade + busca).
   O botÃ£o Exportar Excel usa exatamente o que sobra visÃ­vel aqui. ---- */
function bcmsFiltra(sfx){
  var tid='tab-emtela-'+sfx;
  var cont=document.getElementById(tid); if(!cont) return;
  var tb=cont.querySelector('tbody'); if(!tb) return;
  var v=function(id){var e=document.getElementById(id+'-'+sfx); return e?e.value:'';};
  var q=(v('q-tab-emtela')||'').toLowerCase();
  var fFonte=v('f-fonte'), fAcao=v('f-acao'), fNd=v('f-nd'), fId=v('f-idade');
  var n=0, soma=0;
  tb.querySelectorAll('tr').forEach(function(r){
    var ok=true;
    if(fFonte && r.getAttribute('data-fonte')!==fFonte) ok=false;
    if(ok&&fAcao && r.getAttribute('data-acao')!==fAcao) ok=false;
    if(ok&&fNd && r.getAttribute('data-nd')!==fNd) ok=false;
    if(ok&&fId && r.getAttribute('data-faixa')!==fId) ok=false;
    if(ok&&q && r.textContent.toLowerCase().indexOf(q)===-1) ok=false;
    r.style.display=ok?'':'none';
    if(ok){n++; soma+=parseFloat(r.getAttribute('data-val')||'0')||0;}
  });
  var cnt=document.getElementById('cnt-'+tid);
  if(cnt){var u=cnt.getAttribute('data-unit')||'linha(s)'; cnt.textContent=n+' '+u;}
  /* rodapÃ© passa a refletir o subconjunto filtrado */
  var tf=cont.querySelector('tfoot tr');
  if(tf){
    var tds=tf.querySelectorAll('td');
    var ativo=!!(fFonte||fAcao||fNd||fId||q);
    if(tds.length){tds[0].textContent=(ativo?'FILTRADO Â· ':'TOTAL Â· ')+n+' Nota(s) de CrÃ©dito em tela';}
    if(tds.length>1){tds[tds.length-2].textContent=bcmsBRL(soma);}
  }
  var res=document.getElementById('flt-res-'+sfx);
  if(res){
    var pk=[];
    if(fFonte)pk.push('Fonte '+fFonte);
    if(fAcao)pk.push('AÃ§Ã£o '+fAcao);
    if(fNd)pk.push('ND '+fNd);
    if(fId)pk.push({v:'â‰¤30d',a:'31â€“60d',r:'>60d'}[fId]||fId);
    if(q)pk.push('â€œ'+q+'â€');
    res.textContent=pk.length?(pk.join(' Â· ')+' â€” '+bcmsBRL(soma)):'';
    res.className='flt-resumo'+(pk.length?' on':'');
  }
}
function bcmsLimpaFiltros(sfx){
  ['f-fonte','f-acao','f-nd','f-idade'].forEach(function(id){
    var e=document.getElementById(id+'-'+sfx); if(e)e.value='';});
  var b=document.getElementById('q-tab-emtela-'+sfx); if(b)b.value='';
  bcmsFiltra(sfx);
}
/* Descreve os filtros ativos, p/ registrar no cabeÃ§alho da planilha exportada. */
function bcmsFiltrosAtivos(tid){
  var m=/^tab-emtela-(.+)$/.exec(tid||''); if(!m) return '';
  var sfx=m[1];
  var v=function(id){var e=document.getElementById(id+'-'+sfx); return e?e.value:'';};
  var pk=[];
  if(v('f-fonte'))pk.push('Fonte: '+v('f-fonte'));
  if(v('f-acao'))pk.push('AÃ§Ã£o: '+v('f-acao'));
  if(v('f-nd'))pk.push('ND: '+v('f-nd'));
  var fi=v('f-idade'); if(fi)pk.push('Idade: '+({v:'atÃ© 30 dias',a:'31 a 60 dias',r:'mais de 60 dias'}[fi]||fi));
  var q=v('q-tab-emtela'); if(q)pk.push('Busca: "'+q+'"');
  return pk.length?pk.join(' Â· '):'';
}

function bcmsExportTable(btn,tid,filename){
  var container=document.getElementById(tid);if(!container)return;
  var table=container.tagName==='TABLE'?container:container.querySelector('table');if(!table)return;
  filename=(filename||'creditos_em_tela')+'_'+(new Date().toISOString().slice(0,10));
  var ths=table.querySelectorAll('thead th');var headers=[];var colTypes=[];
  ths.forEach(function(th){
    var txt=th.textContent.replace('â–²','').replace('â–¼','').trim();
    if(txt&&txt!=='AÃ§Ã£o'&&txt!=='AÃ‡ÃƒO'){
      headers.push(txt);
      var u=txt.toUpperCase();
      if(u.indexOf('NOTA')>-1||u.indexOf('NC')>-1||u.indexOf('FONTE')>-1||u.indexOf('UASG')>-1||
         u.indexOf('RECEBIDO EM')>-1||u.indexOf('DATA')>-1||u.indexOf('EMISSÃƒO')>-1||u.indexOf('EMISSAO')>-1||
         u.indexOf('DESCRIÃ‡ÃƒO')>-1||u.indexOf('DESCRICAO')>-1||u.indexOf('OBJETO')>-1||u.indexOf('AÃ‡ÃƒO')>-1||
         u.indexOf('ACAO')>-1||u.indexOf('ND')>-1||u.indexOf('PI')>-1||u.indexOf('OPERAÃ‡ÃƒO')>-1||
         u.indexOf('OPERACAO')>-1||u.indexOf('EMITENTE')>-1||u.indexOf('ORGANIZAÃ‡ÃƒO')>-1||u.indexOf('OMDS')>-1||
         u.indexOf('DIA ANTERIOR')>-1){
        colTypes.push('String');
      } else if(u.indexOf('DIA')>-1||u.indexOf('IDADE')>-1||u.indexOf('CÃ‰LULA')>-1||u.indexOf('CELULA')>-1||u.indexOf('NÂº')>-1||u.indexOf('POS')>-1||u.indexOf('RANK')>-1){
        colTypes.push('Integer');
      } else if(u.indexOf('%')>-1||u.indexOf('TAXA')>-1){
        colTypes.push('Percent');
      } else if(u.indexOf('R$')>-1||u.indexOf('VALOR')>-1||u.indexOf('CRÃ‰DITO')>-1||u.indexOf('CREDITO')>-1||
                 u.indexOf('EMPENHADO')>-1||u.indexOf('LIQUIDADO')>-1||u.indexOf('PAGO')>-1||
                 u.indexOf('PROVISÃƒO')>-1||u.indexOf('PROVISAO')>-1||u.indexOf('SALDO')>-1||
                 u.indexOf('REDUÃ‡')>-1||u.indexOf('LÃQUIDO')>-1||u.indexOf('LIQUIDO')>-1||
                 u.indexOf('RECEBIDO')>-1||u.indexOf('DISP')>-1){
        colTypes.push('Currency');
      } else {
        colTypes.push('String');
      }
    }
  });
  var trs=table.querySelectorAll('tbody tr');var rows=[];
  trs.forEach(function(tr){if(tr.style.display==='none')return;
    var cells=tr.querySelectorAll('td');var rowData=[];
    cells.forEach(function(td,idx){if(idx>=headers.length)return;
      var sortVal=td.getAttribute('data-sort');var exp=colTypes[idx]||'String';
      /* guarda: conteÃºdo com cara de data (dd/mm/aa) Ã© texto, nunca nÃºmero â€”
         senÃ£o a coluna "Recebido" (data) era somada como se fosse moeda. */
      var _txtCel=td.textContent.trim();
      if(/^\d{1,2}\/\d{1,2}\/\d{2,4}$/.test(_txtCel)) exp='String';
      var fullText=td.getAttribute('data-full-desc')||td.getAttribute('title')||td.textContent.replace('â€º','').trim();
      if(!td.getAttribute('data-full-desc')&&!td.getAttribute('title')){fullText=td.textContent.replace('â€º','').trim();}
      if(exp==='String'){
        rowData.push({v:fullText,t:'String',s:'Default'});
      } else if(exp==='Integer'){
        var numVal=sortVal!==null&&!isNaN(parseFloat(sortVal))?parseInt(sortVal,10):parseInt(fullText.replace(/\D/g,''),10);
        if(isNaN(numVal)||numVal<0){rowData.push({v:fullText||'â€”',t:'String',s:'Default'});}
        else{rowData.push({v:numVal,t:'Number',s:'Integer'});}
      } else if(exp==='Currency'){
        var numVal=sortVal!==null&&!isNaN(parseFloat(sortVal))?parseFloat(sortVal):parseFloat(fullText.replace(/[^\d,-]/g,'').replace(',','.'));
        rowData.push({v:isNaN(numVal)?0.0:numVal,t:'Number',s:'Currency'});
      } else if(exp==='Percent'){
        var numVal=sortVal!==null&&!isNaN(parseFloat(sortVal))?parseFloat(sortVal):parseFloat(fullText.replace(/[^\d,-]/g,'').replace(',','.'));
        rowData.push({v:isNaN(numVal)?0.0:numVal/100.0,t:'Number',s:'Percent'});
      } else {
        rowData.push({v:fullText,t:'String',s:'Default'});
      }
    });
    if(rowData.length)rows.push(rowData);
  });
  var xml='<?xml version="1.0" encoding="UTF-8"?>\n'+
   '<?mso-application progid="Excel.Sheet"?>\n'+
   '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"\n'+
   ' xmlns:o="urn:schemas-microsoft-com:office:office"\n'+
   ' xmlns:x="urn:schemas-microsoft-com:office:excel"\n'+
   ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">\n'+
   '<Styles>\n'+
   ' <Style ss:ID="Default" ss:Name="Normal"><Font ss:FontName="Calibri" ss:Size="11" ss:Color="#000000"/><Alignment ss:Vertical="Center"/></Style>\n'+
   ' <Style ss:ID="Header"><Font ss:FontName="Calibri" ss:Size="11" ss:Bold="1" ss:Color="#FFFFFF"/><Interior ss:Color="#1C4A73" ss:Pattern="Solid"/><Alignment ss:Horizontal="Center" ss:Vertical="Center"/></Style>\n'+
   ' <Style ss:ID="Title"><Font ss:FontName="Calibri" ss:Size="14" ss:Bold="1" ss:Color="#1C4A73"/><Alignment ss:Vertical="Center"/></Style>\n'+
   ' <Style ss:ID="Currency"><NumberFormat ss:Format="&quot;R$&quot; #,##0.00"/><Alignment ss:Horizontal="Right" ss:Vertical="Center"/></Style>\n'+
   ' <Style ss:ID="Integer"><NumberFormat ss:Format="#,##0"/><Alignment ss:Horizontal="Center" ss:Vertical="Center"/></Style>\n'+
   ' <Style ss:ID="Percent"><NumberFormat ss:Format="0.0%"/><Alignment ss:Horizontal="Right" ss:Vertical="Center"/></Style>\n'+
   ' <Style ss:ID="Bold"><Font ss:FontName="Calibri" ss:Bold="1"/></Style>\n'+
   '</Styles>\n'+
   '<Worksheet ss:Name="CrÃ©ditos em Tela">\n'+
   '<Table ss:DefaultRowHeight="20">\n';
  headers.forEach(function(h){
    var u=h.toUpperCase();var w=120;
    if(u.indexOf('DESCRIÃ‡ÃƒO')>-1||u.indexOf('DESCRICAO')>-1||u.indexOf('OBJETO')>-1||u.indexOf('APLICAÃ‡ÃƒO')>-1){w=380;}
    else if(u.indexOf('ORGANIZAÃ‡ÃƒO')>-1||u.indexOf('NOME')>-1){w=220;}
    else if(u.indexOf('NC')>-1||u.indexOf('NOTA')>-1){w=140;}
    else if(u.indexOf('DIA')>-1||u.indexOf('IDADE')>-1){w=100;}
    xml+=' <Column ss:AutoFitWidth="1" ss:Width="'+w+'"/>\n';
  });
  xml+=' <Row ss:Height="26"><Cell ss:StyleID="Title" ss:MergeAcross="'+(headers.length-1)+'"><Data ss:Type="String">BASE DE APOIO LOGÃSTICO DO EXÃ‰RCITO â€” RELATÃ“RIO DE CRÃ‰DITOS DISPONÃVEIS</Data></Cell></Row>\n';
  var _flt=(typeof bcmsFiltrosAtivos==='function')?bcmsFiltrosAtivos(tid):'';
  xml+=' <Row ss:Height="18"><Cell ss:MergeAcross="'+(headers.length-1)+'"><Data ss:Type="String">PosiÃ§Ã£o extraÃ­da do Tesouro Gerencial / SIAFI Â· '+new Date().toLocaleDateString('pt-BR')+' Â· '+rows.length+' registro(s)'+(_flt?' Â· FILTROS APLICADOS â€” '+_flt:' Â· sem filtros (lista completa)')+'</Data></Cell></Row>\n';
  xml+=' <Row ss:Height="10"/>\n';
  xml+=' <Row ss:Height="24">\n';
  headers.forEach(function(h){xml+='  <Cell ss:StyleID="Header"><Data ss:Type="String">'+bcmsEscXml(h)+'</Data></Cell>\n';});
  xml+=' </Row>\n';
  rows.forEach(function(r){xml+=' <Row ss:Height="20">\n';
    r.forEach(function(c){
      if(c.t==='Number'){xml+='  <Cell ss:StyleID="'+(c.s||'Currency')+'"><Data ss:Type="Number">'+c.v+'</Data></Cell>\n';}
      else{xml+='  <Cell><Data ss:Type="String">'+bcmsEscXml(c.v)+'</Data></Cell>\n';}
    });
    xml+=' </Row>\n';
  });
  xml+='</Table></Worksheet></Workbook>';
  var blob=new Blob([xml],{type:'application/vnd.ms-excel;charset=utf-8;'});
  var link=document.createElement('a');link.href=URL.createObjectURL(blob);
  link.download=filename+'.xls';document.body.appendChild(link);link.click();document.body.removeChild(link);
  bcmsToast('ðŸ“Š Planilha Excel gerada com sucesso! Download iniciado.');
}

/* ---- Detalhamento por NC "em tela", dividido nas ETAPAS do crÃ©dito ----
   Card principal = SÃ“ o que estÃ¡ realmente disponÃ­vel em tela desta NC.
   Abas: Em tela | HistÃ³rico do PI | LiquidaÃ§Ã£o | Pagamento.
   (Antes o clique abria a cÃ©lula inteira, misturando NCs de objetos diferentes
    que sÃ³ compartilham o mesmo PI.) */
var BTELA=null;
function bcmsTela(row){
  var t=TELADATA[row.getAttribute('data-tela')];
  if(!t){return bcmsCel(row);}   /* fallback: comportamento antigo */
  BTELA=t;
  bcmsTelaAba('tela');
  var m=document.getElementById('modal');
  m.classList.add('open');m.setAttribute('aria-hidden','false');
  var x=document.querySelector('.modal-x');if(x)x.focus();
}
function bcmsTelaAba(aba){
  var t=BTELA;if(!t)return;
  var c=CELDATA[t.cid]||{};
  var fonte=t.u==='OGU'?'OGU (OrÃ§amento Geral da UniÃ£o)':(t.u==='FEx'?'FEx (Fundo do ExÃ©rcito)':(t.u||'â€”'));
  var ncCurta=(String(t.nc).match(/NC(\d+)$/)||[])[1];
  var h='<h3 id="modal-title">'+(ncCurta?'NC '+bcmsEsc(ncCurta):bcmsEsc(t.nc))+'</h3>';
  h+='<p class="m-sub">'+bcmsEsc(t.uasg+' Â· '+fonte+' Â· AÃ§Ã£o '+t.acao+' Â· PI '+t.pi+' Â· ND '+t.nd+(t.ndnome?' â€” '+t.ndnome:''))+'</p>';
  /* etapas */
  var abas=[['tela','ðŸŸ¢ Em tela'],['hist','ðŸ“œ HistÃ³rico do PI'],['liq','ðŸ§¾ LiquidaÃ§Ã£o'],['pag','ðŸ’° Pagamento']];
  h+='<div class="m-etapas" role="tablist">';
  abas.forEach(function(a){
    h+='<button class="m-etapa'+(a[0]===aba?' on':'')+'" role="tab" aria-selected="'+(a[0]===aba)+'" onclick="bcmsTelaAba(\''+a[0]+'\')">'+a[1]+'</button>';
  });
  h+='</div><div class="m-etapa-body">';
  if(aba==='tela'){
    var idade=(t.dias===null||t.dias===undefined)?'â€”':(t.dias+' dia'+(t.dias===1?'':'s'));
    h+='<div class="m-tela-card"><span class="m-tela-lbl">DisponÃ­vel em tela nesta NC</span>'
      +'<b class="m-tela-val">'+bcmsBRL(t.v)+'</b>'
      +'<span class="m-tela-meta">Recebido em '+bcmsEsc(t.dia||'â€”')+' Â· hÃ¡ '+idade+(t.emit?' Â· Emitente '+bcmsEsc(t.emit):'')+'</span></div>';
    if(t.org){
      h+='<div class="m-org"><div class="m-org-h">â†ª CrÃ©dito recebido por <b>mudanÃ§a de ND</b>'
        +(t.nd_de?' (de '+bcmsEsc(t.nd_de)+' para '+bcmsEsc(t.nd)+')':'')+'</div>'
        +'<div class="m-org-b"><span>Objeto original â€” NC '+bcmsEsc(t.org.nc)
        +(t.org.dia?' de '+bcmsEsc(t.org.dia):'')+(t.org.emit?' Â· emitente '+bcmsEsc(t.org.emit):'')+'</span>'
        +'<p>'+bcmsEsc(t.org.obj||'â€”')+'</p></div></div>';
    }
    h+='<div class="m-ncs-h">DescriÃ§Ã£o desta NC</div><div class="m-nc-desc big">'+bcmsEsc(t.obj||'â€”')+'</div>';
    if(t.op) h+='<p class="m-formula">OperaÃ§Ã£o: '+bcmsEsc(t.op)+'</p>';
  } else if(aba==='hist'){
    h+='<p class="m-formula">Todas as movimentaÃ§Ãµes da cÃ©lula <b>'+bcmsEsc(t.acao+' Â· PI '+t.pi+' Â· ND '+t.nd)+'</b> â€” recebimentos, detalhamentos, mudanÃ§as de ND e anulaÃ§Ãµes.</p>';
    h+='<div class="m-kpis"><span>Recebido (lÃ­q)<b>'+bcmsBRL(c.r||0)+'</b></span><span>Empenhado<b>'+bcmsBRL(c.e||0)+'</b></span><span class="ok">DisponÃ­vel<b>'+bcmsBRL(c.d||0)+'</b></span></div>';
    var itens='';
    (c.ncs||[]).forEach(function(n){
      if(!n[0])return;var neg=n[2]<0;var meta=[];
      if(n[4])meta.push('Emitente '+bcmsEsc(n[4]));
      if(n[1])meta.push(bcmsEsc(n[1]));
      if(n[5])meta.push('em '+bcmsEsc(n[5]));
      var ehEsta=(n[0]===t.nc);
      itens+='<div class="m-nc'+(ehEsta?' m-nc-esta':'')+'"><div class="m-nc-h"><span class="m-nc-num">'+bcmsEsc(n[0])+(ehEsta?' <i class="m-nc-tag">esta NC</i>':'')+'</span><span class="m-nc-val'+(neg?' neg':'')+'">'+bcmsBRL(n[2])+'</span></div>';
      if(meta.length)itens+='<div class="m-nc-op">'+meta.join(' Â· ')+'</div>';
      if(n[3])itens+='<div class="m-nc-desc">'+bcmsEsc(n[3])+'</div>';
      itens+='</div>';
    });
    h+='<div class="m-ncs">'+(itens||'<p class="vazio">Sem movimentaÃ§Ãµes.</p>')+'</div>';
  } else if(aba==='liq'){
    var emp=c.e||0,liq=c.l||0;
    h+='<div class="m-kpis"><span>Empenhado<b>'+bcmsBRL(emp)+'</b></span><span class="ok">Liquidado<b>'+bcmsBRL(liq)+'</b></span><span>A liquidar<b>'+bcmsBRL(Math.max(0,emp-liq))+'</b></span></div>';
    h+=bcmsBarra(liq,emp,'Liquidado sobre o empenhado');
    h+='<p class="m-formula">LiquidaÃ§Ã£o Ã© a etapa em que a despesa Ã© atestada (bem/serviÃ§o entregue). Valores da cÃ©lula <b>'+bcmsEsc(t.acao+' Â· PI '+t.pi+' Â· ND '+t.nd)+'</b> â€” o SIAFI nÃ£o segrega liquidaÃ§Ã£o por NC individual.</p>';
  } else {
    var liq2=c.l||0,pag=c.p||0;
    h+='<div class="m-kpis"><span>Liquidado<b>'+bcmsBRL(liq2)+'</b></span><span class="ok">Pago<b>'+bcmsBRL(pag)+'</b></span><span>A pagar<b>'+bcmsBRL(Math.max(0,liq2-pag))+'</b></span></div>';
    h+=bcmsBarra(pag,liq2,'Pago sobre o liquidado');
    h+='<p class="m-formula">Pagamento Ã© a quitaÃ§Ã£o efetiva da despesa liquidada. Valores da cÃ©lula <b>'+bcmsEsc(t.acao+' Â· PI '+t.pi+' Â· ND '+t.nd)+'</b> â€” o SIAFI nÃ£o segrega pagamento por NC individual.</p>';
  }
  h+='</div>';
  document.getElementById('modal-body').innerHTML=h;
}
function bcmsBarra(v,total,rot){
  var p=total>0?Math.min(100,Math.max(0,v/total*100)):0;
  return '<div class="m-barra"><div class="m-barra-t"><span>'+bcmsEsc(rot)+'</span><b>'+p.toFixed(1)+'%</b></div>'
    +'<div class="m-barra-track"><div class="m-barra-fill" style="width:'+p.toFixed(1)+'%"></div></div></div>';
}

function bcmsCel(row){
  var d=CELDATA[row.getAttribute('data-cel')];if(!d)return;
  var h='<h3 id="modal-title">'+bcmsEsc(d.t)+'</h3>';
  var fonte=d.u==='OGU'?'OGU (OrÃ§amento Geral da UniÃ£o)':(d.u==='FEx'?'FEx (Fundo do ExÃ©rcito)':(d.u||'â€”'));
  h+='<div class="m-ficha">'
    +'<span>UASG (Executora)<b>'+bcmsEsc(d.uasg||'â€”')+'</b></span>'
    +'<span>Fonte<b>'+bcmsEsc(fonte)+'</b></span>'
    +'<span>AÃ§Ã£o Governo<b>'+bcmsEsc(d.acao||'â€”')+'</b></span>'
    +'<span class="wide">PI (Plano Interno)<b>'+bcmsEsc(d.pi||'â€”')+(d.pinome?' â€” '+bcmsEsc(d.pinome):'')+'</b></span>'
    +'<span class="wide">ND (Natureza de Despesa)<b>'+bcmsEsc(d.nd||'â€”')+(d.ndnome?' â€” '+bcmsEsc(d.ndnome):'')+'</b></span>'
    +'</div>';
  h+='<div class="m-kpis"><span>Recebido (lÃ­q)<b>'+bcmsBRL(d.r)+'</b></span><span>Empenhado<b>'+bcmsBRL(d.e)+'</b></span><span>Liquidado<b>'+bcmsBRL(d.l||0)+'</b></span><span>Pago<b>'+bcmsBRL(d.p||0)+'</b></span><span class="ok">CrÃ©dito DisponÃ­vel<b>'+bcmsBRL(d.d)+'</b></span></div>';
  h+='<p class="m-formula">Recebido (lÃ­q) âˆ’ Empenhado = CrÃ©dito DisponÃ­vel Â· Empenhado â‰¥ Liquidado â‰¥ Pago</p>';
  var itens='';
  d.ncs.forEach(function(n){
    if(!n[0])return;
    var neg=n[2]<0;
    var meta=[];
    if(n[4]) meta.push('Emitente '+bcmsEsc(n[4]));
    if(n[1]) meta.push(bcmsEsc(n[1]));
    if(n[5]) meta.push('em '+bcmsEsc(n[5]));
    itens+='<div class="m-nc"><div class="m-nc-h"><span class="m-nc-num">'+bcmsEsc(n[0])+'</span><span class="m-nc-val'+(neg?' neg':'')+'">'+bcmsBRL(n[2])+'</span></div>';
    if(meta.length) itens+='<div class="m-nc-op">'+meta.join(' Â· ')+'</div>';
    if(n[3]) itens+='<div class="m-nc-desc">'+bcmsEsc(n[3])+'</div>';
    itens+='</div>';
  });
  var nq=d.ncs.filter(function(n){return n[0];}).length;
  h+='<div class="m-ncs-h">Notas de crÃ©dito da cÃ©lula ('+nq+')</div>';
  h+='<div class="m-ncs">'+(itens||'<p class="vazio">Sem notas de crÃ©dito para detalhar.</p>')+'</div>';
  document.getElementById('modal-body').innerHTML=h;
  var m=document.getElementById('modal');
  m.classList.add('open');
  m.setAttribute('aria-hidden','false');
  var x=document.querySelector('.modal-x');
  if(x) x.focus();
}

function bcmsDay(row){
  var d=DAYDATA[row.getAttribute('data-day')];if(!d)return;
  var h='<h3 id="modal-title">MovimentaÃ§Ã£o de '+bcmsEsc(d.d)+'</h3>';
  h+='<div class="m-kpis"><span>NÂº de NC<b>'+d.n+'</b></span><span>Recebido<b class="col-pos">'+bcmsBRL(d.rec)+'</b></span><span>ReduÃ§Ãµes<b class="col-neg">'+bcmsBRL(d.red)+'</b></span><span class="ok">LÃ­quido<b>'+bcmsBRL(d.liq)+'</b></span></div>';
  h+='<div class="m-ncs-h">Notas de crÃ©dito do dia ('+d.ncs.length+')</div>';
  var itens='';
  d.ncs.forEach(function(n){
    var neg=n[3]<0;
    itens+='<div class="m-nc"><div class="m-nc-h"><span class="m-nc-num">'+bcmsEsc(n[0])+' <span class="pill-fonte">'+bcmsEsc(n[1])+'</span></span><span class="m-nc-val'+(neg?' neg':'')+'">'+bcmsBRL(n[3])+'</span></div>';
    if(n[2]) itens+='<div class="m-nc-op">'+bcmsEsc(n[2])+'</div>';
    if(n[4]) itens+='<div class="m-nc-desc">'+bcmsEsc(n[4])+'</div>';
    itens+='</div>';
  });
  h+='<div class="m-ncs">'+(itens||'<p class="vazio">Sem NC neste dia.</p>')+'</div>';
  document.getElementById('modal-body').innerHTML=h;
  var m=document.getElementById('modal');
  m.classList.add('open');
  m.setAttribute('aria-hidden','false');
  var x=document.querySelector('.modal-x');
  if(x) x.focus();
}

function bcmsCelClose(){
  var m=document.getElementById('modal');
  if(m){ m.classList.remove('open'); m.setAttribute('aria-hidden','true'); }
}

document.addEventListener('keydown', function(e){
  if(e.key==='Escape') bcmsCelClose();
});
"""

# ---------------- main ----------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--local", help="caminho de um xlsx local (teste)")
    ap.add_argument("--date", help="data do snapshot YYYY-MM-DD (default: hoje)")
    ap.add_argument("--file-id", default=(os.environ.get("DRIVE_FILE_ID") or DEFAULT_FILE_ID))
    args = ap.parse_args()

    data_str = args.date or datetime.date.today().isoformat()
    path = args.local if args.local else baixar(args.file_id)
    print("Fonte:", path)
    res, periodo, alertas = etl(path)
    for a in alertas:
        print("[ALERTA]", a)
    hist = atualizar_historico(res, data_str)
    html_out = montar_pagina(res, hist, data_str, periodo, alertas)

    os.makedirs(SITE, exist_ok=True)
    os.makedirs(os.path.join(SITE, "data"), exist_ok=True)
    src_logos = os.path.join(HERE, "assets", "logos")
    dst_logos = os.path.join(SITE, "assets", "logos")
    os.makedirs(dst_logos, exist_ok=True)
    if os.path.isdir(src_logos):
        for fn in os.listdir(src_logos):
            if fn.lower().endswith((".png", ".jpg", ".jpeg", ".webp", ".svg")):
                try:
                    shutil.copyfile(os.path.join(src_logos, fn), os.path.join(dst_logos, fn))
                except Exception as e:
                    print("[AVISO] nÃ£o copiei", fn, ":", e)
    with open(os.path.join(SITE, "index.html"), "w", encoding="utf-8") as f:
        f.write(html_out)
    with open(os.path.join(SITE, "data", "history.json"), "w", encoding="utf-8") as f:
        json.dump(hist, f, ensure_ascii=False, indent=1)

    tot = {k: sum(res[c][k] for c, _ in ALVOS) for k in ("prov", "cred", "emp", "liq", "pag")}
    print(f"OK -> {os.path.join(SITE,'index.html')}")
    print(f"Periodo={periodo} | Credito Disp total={brl(tot['cred'])} | historico {len(hist)} dia(s)")

if __name__ == "__main__":
    main()
