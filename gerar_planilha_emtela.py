# -*- coding: utf-8 -*-
"""
Gera uma planilha .xlsx dos CRÉDITOS EM TELA do BCMS (UASG 160329 e 167329):
Nota de Crédito, objeto da NC, valor em tela e há quantos dias está em tela.

O saldo disponível de cada célula (Ação·PI·ND) é atribuído às Notas de Crédito de
recebimento mais recentes (LIFO) — assim cada linha carrega o objeto real da NC,
o valor que ainda está em tela e a data/idade correspondentes. A soma fecha com o
Crédito Disponível total.

Reaproveita a ETL do dashboard (gerar_dashboard.py).

Uso:
    py -3 gerar_planilha_emtela.py                 # baixa do Drive
    py -3 gerar_planilha_emtela.py --local X.xlsx  # arquivo local
"""
import os, argparse, datetime, json, re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from gerar_dashboard import etl, baixar, UNIDADES, _par, FONTE_CURTA, DEFAULT_FILE_ID, brl

# Esta planilha é específica do BCMS (UASG 160329 e 167329), conforme especificado.
# (No dashboard, ALVOS cobre as 6 OMDS; aqui restringimos ao BCMS.)
ALVOS = _par(UNIDADES[0])

HERE = os.path.dirname(os.path.abspath(__file__))
OUTDIR = os.path.join(HERE, "relatorios")
OUT = os.path.join(OUTDIR, "CREDITOS_EM_TELA_BCMS.xlsx")
RESUMOS_PATH = os.path.join(HERE, "data", "resumos_nc.json")

# Cache de resumos (finalidade + prazo) por número de NC — mantido à parte; NCs novas são reportadas.
try:
    with open(RESUMOS_PATH, encoding="utf-8") as _f:
        RESUMOS = json.load(_f)
except Exception:
    RESUMOS = {}

def extrai_prazo(obj):
    """Fallback p/ NCs sem resumo: tenta extrair o prazo de empenho do texto."""
    o = (obj or "").upper()
    if re.search(r"IMEDIAT|EMPH?\s*IMTO|EMP\s*IMTO", o):
        return "Imediato"
    m = re.search(r"(?:EMPENH\w*|EMPH?)\s*AT[EÉ]\s*:?\s*(\d{1,2}\s*[/ ]?\s*[A-Z]{3,4}\s*[/ ]?\s*\d{2,4}|\d{1,2}/\d{1,2}/\d{2,4})", o)
    if m:
        return re.sub(r"\s+", "", m.group(1))
    m = re.search(r"PRAZO\s*(?:PARA|DE)?\s*EMP\w*\s*:?\s*(\d+\s*DIAS[^.\-]*)", o)
    if m:
        return m.group(1).strip().title()
    return "—"

def resumir(nc, obj):
    r = RESUMOS.get(nc)
    if r:
        return r.get("finalidade", ""), r.get("prazo", "—")
    fin = (obj or "").strip()
    fin = (fin[:70] + "…") if len(fin) > 72 else fin
    return "(resumir) " + fin, extrai_prazo(obj)

def parse_dia(s):
    try:
        dd, mm, yy = str(s).split("/")
        return datetime.date(int(yy), int(mm), int(dd))
    except Exception:
        return None

def build_rows(res, asof):
    rows = []
    for cod, _ in ALVOS:
        d = res[cod]
        cel_recs = {}
        for L in d["linhas"]:
            if L["cred"] > 0 and L.get("dia"):
                cel_recs.setdefault((L["acao"], L["pi"], L["nd"]), []).append(L)
        for cl in d["celulas"].values():
            if cl["cred"] <= 0.005:
                continue
            recs = sorted(cel_recs.get((cl["acao"], cl["pi"], cl["nd"]), []),
                          key=lambda L: parse_dia(L["dia"]) or datetime.date.min, reverse=True)
            uasg_lbl = f"{cod} ({FONTE_CURTA[cod]})"
            pi_ext = cl.get("pi_nome") or cl["pi"] or "(sem PI)"
            restante = cl["cred"]
            for L in recs:
                if restante <= 0.005:
                    break
                val = min(L["cred"], restante)
                restante -= val
                dt = parse_dia(L["dia"])
                fin, prz = resumir(L["nc"], L["obj"])
                rows.append(dict(uasg=cod, uasg_lbl=uasg_lbl, pi_ext=pi_ext, acao=cl["acao"], pi=cl["pi"],
                                 nd=cl["nd"], nd_nome=cl.get("nd_nome", ""), nc=L["nc"], obj=L["obj"],
                                 op=L["op"], data=L["dia"], dt=dt, finalidade=fin, prazo=prz,
                                 dias=((asof - dt).days if dt else None), valor=val))
            if restante > 0.005:  # saldo sem NC datada — não perder o valor
                rows.append(dict(uasg=cod, uasg_lbl=uasg_lbl, pi_ext=pi_ext, acao=cl["acao"], pi=cl["pi"],
                                 nd=cl["nd"], nd_nome=cl.get("nd_nome", ""), nc="(sem NC datada)",
                                 obj="(saldo sem NC datada)", op="", data="", dt=None,
                                 finalidade="(saldo sem NC datada)", prazo="—", dias=None, valor=restante))
    rows.sort(key=lambda r: (r["dias"] if r["dias"] is not None else -1), reverse=True)
    return rows

# estilos
AZUL, VERDE, CINZA = "1F4E78", "375623", "D9D9D9"
f_tit = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
f_sub = Font(name="Calibri", size=10, color="FFFFFF")
f_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
f_b = Font(name="Calibri", size=10, bold=True)
f_n = Font(name="Calibri", size=10)
fill_tit = PatternFill("solid", fgColor=AZUL)
fill_sub = PatternFill("solid", fgColor="2E75B6")
fill_hdr = PatternFill("solid", fgColor="2E75B6")
fill_tot = PatternFill("solid", fgColor=CINZA)
thin = Side(style="thin", color="BFBFBF")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '#,##0.00'
al_l = Alignment(horizontal="left", vertical="center")
al_c = Alignment(horizontal="center", vertical="center")
al_r = Alignment(horizontal="right", vertical="center")
al_w = Alignment(horizontal="left", vertical="top", wrap_text=True)
VERM = Font(name="Calibri", size=10, bold=True, color="C00000")
AMB = Font(name="Calibri", size=10, bold=True, color="BF8F00")
VER = Font(name="Calibri", size=10, color="375623")

COLS = [("UASG", "uasg_lbl", 15, al_c),
        ("PI (por extenso)", "pi_ext", 30, al_w),
        ("Finalidade", "finalidade", 54, al_w),
        ("Prazo p/ empenhar", "prazo", 16, al_c),
        ("Valor (R$)", "valor", 15, al_r),
        ("Dias em tela", "dias", 12, al_c)]

def gerar(res, periodo, rows):
    total = sum(r["valor"] for r in rows)
    ids = [(r["dias"], r["valor"]) for r in rows if r["dias"] is not None]
    idade_pond = round(sum(d * v for d, v in ids) / sum(v for _, v in ids)) if ids else 0
    mais_antigo = max((d for d, _ in ids), default=0)
    hoje = datetime.date.today().strftime("%d/%m/%Y %H:%M")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Créditos em tela"
    ws.sheet_view.showGridLines = False
    n = len(COLS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(1, 1, "CRÉDITOS EM TELA — BCMS (UASG 160329 e 167329)"); c.font = f_tit; c.alignment = al_l
    for i in range(1, n + 1): ws.cell(1, i).fill = fill_tit
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    c = ws.cell(2, 1, f"Fonte: CRÉDITO DISP (Tesouro Gerencial/SIAFI) · Posição {periodo or '—'} · dados com ~24h de defasagem · gerado em {hoje}")
    c.font = f_sub; c.alignment = al_l
    for i in range(1, n + 1): ws.cell(2, i).fill = fill_sub
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n)
    c = ws.cell(3, 1, f"Total em tela: {brl(total)}  ·  {len(rows)} nota(s) de crédito  ·  idade média (ponderada) {idade_pond} dias  ·  mais antigo {mais_antigo} dias")
    c.font = f_b; c.alignment = al_l

    hr = 5
    for i, (h, k, w, al) in enumerate(COLS, 1):
        cell = ws.cell(hr, i, h); cell.font = f_hdr; cell.fill = fill_hdr; cell.alignment = al_c; cell.border = bd
        ws.column_dimensions[get_column_letter(i)].width = w

    r = hr + 1
    for row in rows:
        for i, (h, k, w, al) in enumerate(COLS, 1):
            v = row[k]
            cell = ws.cell(r, i)
            cell.alignment = al; cell.border = bd; cell.font = f_n
            if k == "valor":
                cell.value = round(v, 2); cell.number_format = MONEY; cell.font = f_b
            elif k == "dias":
                if v is None:
                    cell.value = "—"
                else:
                    cell.value = v
                    cell.font = VERM if v > 60 else (AMB if v > 30 else VER)
            else:
                cell.value = v
        r += 1
    # total (na coluna Valor)
    vcol = [k for _, k, _, _ in COLS].index("valor") + 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=vcol - 1)
    ws.cell(r, 1, "TOTAL").font = f_b; ws.cell(r, 1).alignment = al_c
    for i in range(1, n + 1):
        ws.cell(r, i).fill = fill_tot; ws.cell(r, i).border = bd
    tc = ws.cell(r, vcol, round(total, 2)); tc.font = f_b; tc.number_format = MONEY; tc.alignment = al_r

    ws.freeze_panes = f"A{hr+1}"
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(n)}{r-1}"
    os.makedirs(OUTDIR, exist_ok=True)
    wb.save(OUT)
    return total, idade_pond, mais_antigo

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--local")
    ap.add_argument("--file-id", default=(os.environ.get("DRIVE_FILE_ID") or DEFAULT_FILE_ID))
    args = ap.parse_args()
    path = args.local if args.local else baixar(args.file_id)
    print("Fonte:", path)
    res, periodo, alertas = etl(path)
    for a in alertas: print("[ALERTA]", a)
    # asof = data de emissão de NC mais recente na base (consistente com o painel)
    todas = [parse_dia(L["dia"]) for cod, _ in ALVOS for L in res[cod]["linhas"] if L.get("dia")]
    todas = [d for d in todas if d]
    asof = max(todas) if todas else datetime.date.today()
    rows = build_rows(res, asof)
    faltando = sorted({r["nc"] for r in rows if r["nc"] and not str(r["nc"]).startswith("(") and r["nc"] not in RESUMOS})
    if faltando:
        print(f"[RESUMO] {len(faltando)} NC(s) SEM resumo no cache (finalidade provisória — resumir e adicionar em data/resumos_nc.json):")
        for nc in faltando:
            obj = next((L["obj"] for cod, _ in ALVOS for L in res[cod]["linhas"] if L["nc"] == nc), "")
            print(f"   - {nc}: {obj[:110]}")
    total, media, antigo = gerar(res, periodo, rows)
    print(f"OK -> {OUT}")
    print(f"{len(rows)} NCs | Total em tela {brl(total)} | idade média {media}d | mais antigo {antigo}d | asof {asof}")

if __name__ == "__main__":
    main()
